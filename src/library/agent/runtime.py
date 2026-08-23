"""Agent runtime — DESIGN.md §10.2 + §12.2.

Plan-Execute loop, exposed as async generator yielding AgentEvent frames
for SSE streaming. One `run_turn(session_id, user_message)` invocation:

  1. Open one conversation row (turn_index = next). Yield "conversation".
  2. Plan phase: yield "planning", do ONE LLM call with `tools=[]`,
     yield "plan" with the user-visible plan text. Stored in
     conversations.llm_calls under phase='plan'. If plan_text starts with
     `NO_PLAN:` the trailing answer is treated as the final answer unless the
     current user turn visibly requires Library's local tools.
  3. Execute phase: up to `settings.agent_execute_max_turns` (default 15)
     LLM calls. For each:
         - yield "thinking", LLM call (records usage)
         - if model returned tool_calls: yield "tool_call" per call,
           dispatch (with dedup + doom-loop guards), yield "tool_result",
           feed back as `tool` message
         - if model returned text + no tool_calls AND stop_reason='end_turn':
           yield "answer" with final text
         - if final text hits stop_reason='max_tokens', continue the answer
           server-side and emit one merged "answer" event.
     Once the run enters the last 1/3 of the budget, append wrap-up tail.
  4. Truncation: if the turn budget is hit, yield "answer" with fallback
     text and mark truncated=True.
  5. Finalize: write agent_response, ended_at; enqueue reflect_turn task
     (priority 30); record task_outcome; yield "done" with usage JSON.

Guards (added 2026-05-24, all append-only — never mutate prior messages
so ephemeral cache breakpoints stay valid):
  - NO_PLAN fast-path: planner can opt out of execute for trivial turns; the
    runtime repairs that decision when the current turn visibly needs local
    Library tools.
  - Tool-call dedup: identical (name, args) within one turn returns the
    prior result synthetically without re-dispatching.
  - Doom-loop guard: if the same (name, args) appears K times in the last
    N tool calls, the next tool result message gets a STOP nudge appended.

Concurrency: this runtime assumes one in-flight turn per session. The
HTTP route layer (api/routes_chat.py) enforces this with a per-session
asyncio.Lock held for the whole SSE stream; the cross-process backstop
is `UNIQUE(session_id, turn_index)` on `conversations`. Anything else
calling `run_turn` directly (tests, scripts) MUST serialise per session
or risk duplicate-row IntegrityError on the second writer.
"""
from __future__ import annotations

import asyncio
import copy
import json
import logging
import re
import time
import urllib.parse
from collections import deque
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from typing import Any, AsyncIterator, Literal, Mapping, Sequence

from library.agent.compression_adapter import maybe_compress_tool_result_for_model
from library.agent.cache_metrics import summarize_llm_calls
from library.agent.conversation_compaction import (
    TokenCounter,
    fit_messages_to_token_budget,
)
from library.agent.citation_manifest import (
    attach_citation_manifest,
    prepare_finish_citation_manifest,
)
from library.agent.stable_context import (
    build_plan_history_messages,
    build_resumed_messages,
    build_stable_snapshot,
    build_snapshot_messages,
    render_phase_system_prompt,
)
from library.agent.tool_scheduler import ScheduledTool, schedule_waves
from library.agent.tool_locks import tool_execution_lock
from library.agent.tools import ToolContext, all_tool_defs, get_tool
from library.agent.types import AgentEvent, AgentTurnError, RunOptions, TurnUsage
from library.citations import (
    CITATION_FOOTNOTE_RE,
    CitationFootnote,
    parse_citation_footnote_match,
    quote_matches_source_text,
    unescape_citation_quote,
)
from library.db.models import Conversation as ConversationRow, Session as SessionRow
from library.db.session import session_scope
from library.llm import (
    ChatMessage,
    ChatRequest,
    ContentBlock,
    ImageBlock,
    TextBlock,
    ToolResultBlock,
    ToolUseBlock,
    PromptPrefixTracker,
    get_chat_client,
)
from library.config import get_settings, has_vision_profile
from library.pipelines.pdf_text import (
    PdfTextRange,
    first_page_number,
    get_pdf_page_labels_for_file,
    get_pdf_text_for_file,
    locate_quote_page,
    resolve_page_label,
)
from library.repositories import sessions as session_service
from library.repositories import entries as entries_repo
from library.repositories import tags as tags_repo
from library.repositories import folders as folders_repo
from library.repositories import catalogs as catalogs_repo
from library.repositories.task_outcomes import record_outcome
from library.services.attachments import save_turn_attachments
from library.tasks.enqueue import enqueue
from library.tasks.kinds import KIND_REFLECT_TURN
from library.agent import tool_display

log = logging.getLogger(__name__)

MAX_TOOL_RESULT_LEN = 50_000
QUICK_EXECUTE_MAX_TURNS = 4
STANDARD_EXECUTE_MAX_TURNS = 8
AUTO_MAX_BUDGET_UPGRADES = 2
MALFORMED_TOOL_ARGUMENT_REPAIR_LIMIT = 2
QUICK_FORCED_ANSWER_RETRIES = 1
MAX_FINALIZATION_ATTEMPTS = 2
# Structured-truncation safety net: how many trim passes before falling
# back to string slicing. Practically each pass halves one large list, so
# 3 passes can absorb three different oversize lists in one payload.
STRUCTURED_TRUNCATE_PASSES = 3
# Default token budgets — overridable per-deploy via AGENT_PLAN_MAX_TOKENS /
# AGENT_EXECUTE_MAX_TOKENS in settings. Sized for gpt-4o-class models; bump
# for long-context backends (DeepSeek-V3, Claude 3.5 Sonnet, etc.).
PLAN_MAX_TOKENS = 1024
EXECUTE_MAX_TOKENS = 2048
TOOL_RESULT_PREVIEW_LEN = 240

NO_PLAN_PREFIX = "NO_PLAN:"
BUDGET_PREFIX = "BUDGET:"
SESSION_NAME_PREFIX = "Session name:"
MAX_SESSION_NAME_LEN = 80
BudgetTier = Literal["quick", "standard", "deep"]
AnswerPhase = Literal["researching", "finalizing"]
BUDGET_TIERS: tuple[BudgetTier, ...] = ("quick", "standard", "deep")

# Doom-loop: if the same (name, canonical_args) shows up
# DOOM_LOOP_THRESHOLD times within the last DOOM_LOOP_WINDOW tool calls,
# inject a STOP nudge. The threshold is one above the dedup floor — dedup
# already neutralises duplicate work, so this fires only on near-duplicate
# patterns the model is iterating on (slightly different args each time).
DOOM_LOOP_WINDOW = 6
DOOM_LOOP_THRESHOLD = 3
DOOM_LOOP_NUDGE = (
    "[runtime guard] You have repeatedly called the same tool with similar "
    "arguments. Stop expanding tool calls and give the final answer from the "
    "results already collected."
)
FINAL_ANSWER_CONTINUE_NUDGE = (
    "[runtime guard] Your previous final answer was cut off by the token "
    "limit. Continue exactly where it stopped. Do not restart, do not repeat "
    "previous text, do not call tools, and finish the answer in the same "
    "language as the user's latest original question unless they explicitly asked "
    "otherwise."
)
QUICK_FORCED_ANSWER_NUDGE = (
    "[runtime guard] Your previous response attempted a tool call, but Quick "
    "mode has reached the final answer round and tools are unavailable. "
    "Do not call tools. Do not emit DSML, XML, JSON, or pseudo function-call "
    "markup. Write the final answer now from the evidence already collected. "
    "If evidence is incomplete, state the missing piece and give the best "
    "bounded answer. Use the same language as the user's latest original "
    "question unless they explicitly asked otherwise."
)
PREMATURE_NO_TOOL_NUDGE = (
    "[runtime guard] The research phase is still active, so a text-only response "
    "cannot complete this turn. If evidence is still missing, call the appropriate "
    "retrieval or read tool now. If evidence is sufficient, or targeted checks "
    "established that it is unavailable, call `finish_research` now. Do not "
    "describe a future tool call in text."
)
FINALIZE_RESEARCH_NUDGE = (
    "[research complete] Evidence gathering is closed. Write the complete final "
    "Markdown answer to the user's latest original question now, using only the "
    "evidence already collected. Do not call tools or describe future work. "
    "If finish_research returned a citation_manifest, place its assigned markers "
    "after the corresponding supported claims and do not recreate their footnote "
    "definitions; the runtime appends them deterministically. If no manifest was "
    "returned, follow the ordinary citation contract. Return the entire answer body."
)
FINALIZATION_RETRY_NUDGE = (
    "[runtime guard] The previous finalizing response was not a valid complete "
    "answer: {issue}. Return the entire corrected Markdown answer now. Do not "
    "call tools or omit supported claims."
)
MALFORMED_TOOL_ARGUMENT_NUDGE = (
    "[runtime guard] One or more tool calls were not executed because their "
    "arguments were not valid JSON. Retry the needed call now with exactly one "
    "complete JSON object matching the tool schema. Do not use Markdown code "
    "fences, comments, trailing commas, or blank values."
)

_FUTURE_ACTION_RE = re.compile(
    r"(?:^|[.!?]\s+)(?:let me|i(?:'ll| will)|i need to)\s+"
    r"(?:read|search|look|inspect|check|open|fetch|retrieve|query|call|use)\b"
    r"|(?:让我|我来|接下来我?(?:会|将|要))"
    r"(?:读取|搜索|查找|查看|检查|打开|获取|检索|查询|调用|使用)",
    re.IGNORECASE,
)

_KB_TOOL_HINT_RE = re.compile(
    r"\b("
    r"file|files|document|documents|doc|docs|pdf|note|notes|entry|entries|"
    r"folder|folders|catalog|catalogs|tag|tags|journal|journals|library|"
    r"knowledge[- ]base|uploaded|stored|citation|citations|quote|quotes"
    r")\b",
    re.IGNORECASE,
)
_KB_TOOL_HINT_CJK = (
    "文件",
    "文档",
    "资料",
    "笔记",
    "条目",
    "目录",
    "标签",
    "知识库",
    "库里",
    "上传",
    "材料",
    "引用",
    "证据",
    "来源",
)
_KB_LOCAL_REF_CJK = (
    "这篇",
    "这份",
    "这段",
)
_KB_LOCAL_ACTION_CJK = (
    "总结",
    "概括",
    "分析",
    "阅读",
    "看看",
    "看一下",
    "提炼",
    "引用",
    "证据",
    "来源",
    "内容",
    "里面",
    "讲什么",
)


def _requires_library_tools(user_message: str) -> bool:
    """Heuristic for post-hoc repair when a model skips tools too early.

    The planner is allowed to route weather/small-talk/general out-of-scope
    turns to direct answers. This guard is deliberately narrower: it only
    repairs no-tool execute answers for requests that visibly point at the
    local Library library.
    """
    text = user_message or ""
    if _KB_TOOL_HINT_RE.search(text):
        return True
    if any(term in text for term in _KB_TOOL_HINT_CJK):
        return True
    return (
        any(ref in text for ref in _KB_LOCAL_REF_CJK)
        and any(action in text for action in _KB_LOCAL_ACTION_CJK)
    )


def _no_plan_repair_plan(user_message: str, *, mode: str) -> str:
    """Fallback plan when the planner incorrectly emits NO_PLAN for local data.

    Ambiguous local references should still go through execute: the executor can
    inspect likely local context and ask a focused clarification if it cannot
    identify the target material.
    """
    tier = "deep" if mode == "deep" else "quick"
    if _prefers_zh(user_message):
        return (
            f"BUDGET: {tier}\n"
            "1. 定位用户问题指向的本地资料、笔记或知识库内容。\n"
            "2. 读取关键证据；如果仍无法确定目标材料，给出明确的澄清问题。"
        )
    return (
        f"BUDGET: {tier}\n"
        "1. Identify the local files, notes, or knowledge-base material the "
        "user's turn points to.\n"
        "2. Read the key evidence; if the target remains ambiguous, ask a "
        "focused clarification question."
    )


async def _record_no_plan_repair(
    conversation_id: str,
    *,
    repaired_plan_text: str,
    raw_no_plan_answer: str,
) -> None:
    """Replace the persisted public plan after a planner NO_PLAN repair."""
    try:
        async with session_scope() as db:
            conv = await db.get(ConversationRow, conversation_id)
            if conv is None:
                return
            calls = list(conv.llm_calls or [])
            for idx, call in enumerate(calls):
                if not isinstance(call, dict) or call.get("phase") != "plan":
                    continue
                updated = dict(call)
                raw_plan = updated.get("plan_text")
                if isinstance(raw_plan, str) and raw_plan.strip():
                    updated["raw_plan_text"] = raw_plan
                updated["plan_text"] = repaired_plan_text
                updated["no_plan_repaired"] = True
                updated["raw_no_plan_answer"] = raw_no_plan_answer
                calls[idx] = updated
                conv.llm_calls = calls
                await db.commit()
                return
    except Exception:
        log.exception(
            "failed to record planner NO_PLAN repair for conversation %s",
            conversation_id,
        )


def _canonical_args(arguments: Any) -> str:
    """Stable JSON serialisation of tool arguments for dedup keying.

    `sort_keys=True` so {a:1,b:2} and {b:2,a:1} hash identical; we accept
    that nested-dict ordering still collapses correctly because json.dumps
    recursively sorts. None-valued fields keep their slot — different from
    "field absent" — to avoid false dedup of intentionally-distinct calls.
    """
    try:
        return json.dumps(arguments, ensure_ascii=False, sort_keys=True)
    except (TypeError, ValueError):
        return repr(arguments)


def _find_largest_list(payload: Any) -> tuple[list, str, int] | None:
    """Walk `payload` and return the longest list with its dotted path and
    serialized weight, or None if the payload contains no lists. We pick by
    serialized character cost — a list of 5 huge dicts outranks a list of
    500 ints — because that's what the budget actually constrains."""
    best: tuple[list, str, int] | None = None

    def visit(node: Any, path: str) -> None:
        nonlocal best
        if isinstance(node, list):
            try:
                weight = len(json.dumps(node, ensure_ascii=False))
            except (TypeError, ValueError):
                weight = sum(len(repr(x)) for x in node)
            if best is None or weight > best[2]:
                best = (node, path or "$", weight)
            for i, item in enumerate(node):
                visit(item, f"{path}[{i}]")
        elif isinstance(node, dict):
            for k, v in node.items():
                visit(v, f"{path}.{k}" if path else k)

    visit(payload, "")
    return best


def _trim_largest_list(payload: Any, budget: int) -> tuple[bool, str | None, int]:
    """Find the largest list in `payload` and shrink it (in place) until the
    re-serialized payload fits within `budget`. Returns
    (changed, path, dropped_count). If no list is found or trimming cannot
    bring the payload under budget, returns (False, None, 0)."""
    target = _find_largest_list(payload)
    if target is None:
        return False, None, 0
    lst, path, _ = target
    original = list(lst)
    n = len(original)
    if n == 0:
        return False, None, 0
    # Binary-search the largest prefix that fits. lst is mutated in place,
    # so we save `original` once and restore the prefix on each probe.
    lo, hi = 0, n
    while lo < hi:
        mid = (lo + hi + 1) // 2
        lst.clear()
        lst.extend(original[:mid])
        try:
            size = len(json.dumps(payload, ensure_ascii=False, default=str))
        except (TypeError, ValueError):
            size = budget + 1
        if size <= budget:
            lo = mid
        else:
            hi = mid - 1
    lst.clear()
    lst.extend(original[:lo])
    dropped = n - lo
    return dropped > 0, path, dropped


def _copy_jsonish(value: Any) -> Any:
    """Return a mutation-safe copy for model-only truncation."""
    try:
        return copy.deepcopy(value)
    except Exception:
        try:
            return json.loads(json.dumps(value, ensure_ascii=False, default=str))
        except Exception:
            return value


def _structured_truncate(payload: Any, budget: int) -> tuple[str, dict | None]:
    """Serialize `payload` to JSON ≤ `budget` chars by trimming its largest
    lists. Returns (json_text, marker) where `marker` describes what was
    dropped (or None when nothing was trimmed). Falls back to a string
    slice on the serialized output if structured passes can't shrink it
    enough — that branch should be rare and signals an oddly-shaped
    payload (deeply nested scalars, no lists)."""
    try:
        text = json.dumps(payload, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        return repr(payload)[:budget], None
    if len(text) <= budget:
        return text, None
    if not isinstance(payload, (dict, list)):
        return text[:budget] + "...(truncated)", {
            "_truncated_field": "$", "_truncated_dropped": -1,
            "_truncated_reason": "non-container payload",
        }
    # Reserve headroom for the marker we'll inject after trimming, so the
    # final post-marker payload still fits within `budget`.
    MARKER_HEADROOM = 240
    inner_budget = max(budget - MARKER_HEADROOM, budget // 2)
    truncations: list[dict[str, Any]] = []
    for _ in range(STRUCTURED_TRUNCATE_PASSES):
        changed, path, dropped = _trim_largest_list(payload, inner_budget)
        if not changed:
            break
        truncations.append({"path": path, "dropped": dropped})
        try:
            text = json.dumps(payload, ensure_ascii=False, default=str)
        except (TypeError, ValueError):
            break
        if len(text) <= inner_budget:
            break
    marker: dict[str, Any] = {}
    if truncations:
        first = truncations[0]
        marker["_truncated_field"] = first["path"]
        marker["_truncated_dropped"] = first["dropped"]
        if len(truncations) > 1:
            marker["_truncated_path"] = truncations
    if isinstance(payload, dict) and marker:
        payload.update(marker)
    try:
        text = json.dumps(payload, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        pass
    if len(text) > budget:
        text = text[:budget] + "...(truncated)"
        marker.setdefault("_truncated_reason", "fallback string slice")
    return text, (marker or None)


@dataclass(slots=True)
class _CallGuard:
    """Per-turn tracker for dedup + doom-loop detection."""
    seen: dict[str, str] = field(default_factory=dict)  # key -> prior result_text
    seen_previews: dict[str, str] = field(default_factory=dict)  # key -> user-facing preview
    recent: deque[str] = field(default_factory=lambda: deque(maxlen=DOOM_LOOP_WINDOW))
    nudged: bool = False

    def key(self, name: str, arguments: Any) -> str:
        return f"{name}::{_canonical_args(arguments)}"

    def remember(self, key: str, result_text: str, preview: str = "") -> None:
        self.seen[key] = result_text
        if preview:
            self.seen_previews[key] = preview
        self.recent.append(key)

    def is_duplicate(self, key: str) -> bool:
        return key in self.seen

    def should_nudge(self, key: str) -> bool:
        """True the *first* time the loop pattern crosses threshold.

        We count `key` in the rolling window but don't include the current
        call yet (caller decides whether to record it). Once nudged, never
        nudges again in the same turn — one warning is enough; piling on
        wastes tokens and pollutes the next prefix-cache hit.
        """
        if self.nudged:
            return False
        return self.recent.count(key) + 1 >= DOOM_LOOP_THRESHOLD


@dataclass(slots=True)
class _ExecuteOutcome:
    """Mutable carrier returned by `_run_execute_phase` so the caller can
    pick up the final answer text and truncation flag without needing a
    sentinel event in the public stream."""
    answer: str = ""
    truncated: bool = False
    error: str | None = None


@dataclass(slots=True)
class _DispatchStats:
    """One execute round's useful-work summary for auto-budget routing."""
    successful_new_results: int = 0
    finish_research_requested: bool = False
    evidence_status: str | None = None
    citation_manifest: list[dict[str, Any]] | None = None


@dataclass(slots=True)
class _BudgetState:
    requested_mode: str
    initial_tier: BudgetTier
    current_tier: BudgetTier
    limit: int
    hard_limit: int
    source: str
    max_upgrades: int = 0
    upgrades: int = 0

    @property
    def auto(self) -> bool:
        return self.requested_mode == "auto"

    def payload(self) -> dict[str, Any]:
        return {
            "mode": self.requested_mode,
            "tier": self.current_tier,
            "initial_tier": self.initial_tier,
            "limit": self.limit,
            "hard_limit": self.hard_limit,
            "source": self.source,
            "upgrades": self.upgrades,
        }


# ---- multimodal current-turn helpers --------------------------------------

# Max images we describe via the vision fallback before planning starts.
# Each is one sequential vision call, so this bounds the pre-plan latency
# (agent_turn_timeout_seconds still applies on top).
_VISION_FALLBACK_MAX_IMAGES = 4
# Generous budget: a *reasoning* vision model (e.g. Qwen "thinking" variants)
# spends most of its output tokens on hidden reasoning before emitting any
# visible description, so a small cap gets consumed by reasoning and returns an
# empty (— "unavailable") description. Leave ample room for reasoning + a full
# transcription/description.
_VISION_FALLBACK_MAX_TOKENS = 4096
_VISION_FALLBACK_MARKER = (
    "--- Pasted image (auto-described; chat model has no vision) ---"
)


# --- vision capability probe (chat_vision="auto") --------------------------
# A 1x1 grayscale PNG. Sent as a throwaway image so a text-only provider
# rejects it with a 400 exactly as it would a real image, letting us classify
# the model without spending real vision tokens.
_PROBE_IMAGE = ImageBlock(
    media_type="image/png",
    data_b64=(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAAAAAA6fptVAAAACklEQVR4nGNgAAAA"
        "AgABSK+kcQAAAABJRU5ErkJggg=="
    ),
)
# Per-process cache: (provider, model) -> supports image input. Reset on
# restart; each API/worker process probes a given model at most once.
_vision_capability: dict[tuple[str, str], bool] = {}
def _looks_like_vision_unsupported(exc: Exception) -> bool:
    """A 400 that clearly blames image/vision input (vs an unrelated bad
    request like auth/quota/params). We only degrade to the describe fallback
    on a clear image signal; an ambiguous 400 is left to fail on the real call
    so the user sees the true error rather than silently losing the image."""
    msg = str(exc).lower()
    if any(w in msg for w in ("image", "vision", "multimodal", "image_url")):
        return True
    # Some providers phrase it as a modality/content-type support problem.
    return "modality" in msg and "support" in msg


async def _chat_model_supports_vision() -> bool:
    """Whether the `chat` profile's model accepts image input.

    Probes once per (provider, model) with a 1x1 image and caches the verdict.
    A vision-unsupported 400 => False; success => True. An ambiguous/transient
    error (auth, rate limit, network) is NOT cached and optimistically treated
    as vision-capable so the real turn proceeds (and surfaces that error).
    """
    from openai import BadRequestError as _OpenAIBadRequest
    from anthropic import BadRequestError as _AnthropicBadRequest

    client = get_chat_client("chat")
    capabilities = getattr(client, "capabilities", None)
    if capabilities is not None:
        return bool(capabilities.supports_vision)
    key = (getattr(client, "provider", "?"), getattr(client, "model", "?"))
    cached = _vision_capability.get(key)
    if cached is not None:
        return cached
    try:
        await client.complete(ChatRequest(
            system=None,
            messages=[ChatMessage(
                role="user",
                content=[TextBlock(text="."), _PROBE_IMAGE],
            )],
            max_tokens=1,
        ))
        _vision_capability[key] = True
    except (_OpenAIBadRequest, _AnthropicBadRequest) as exc:
        if _looks_like_vision_unsupported(exc):
            log.info("chat model %s rejected image input; using vision "
                     "describe fallback for pasted images", key)
            _vision_capability[key] = False
        else:
            # Unrelated 400 — don't mislabel the model; let the real call run.
            return True
    except Exception:
        # Transient (rate limit / network); assume capable this time.
        log.warning("vision capability probe for %s failed transiently; "
                    "assuming vision-capable", key, exc_info=True)
        return True
    return _vision_capability[key]


def _current_user_content(
    user_message: str,
    images: list[ImageBlock] | None,
) -> str | list[ContentBlock]:
    """Build the current turn's user message content.

    Plain string when there are no images (keeps history/replay text-only by
    construction). When images are present, returns a block list with the
    ImageBlocks attached to THIS turn only; the leading TextBlock is omitted
    for an image-only turn (blank caption).
    """
    if not images:
        return user_message
    blocks: list[ContentBlock] = []
    if user_message.strip():
        blocks.append(TextBlock(text=user_message))
    blocks.extend(images)
    return blocks


def _persisted_user_message(user_message: str, image_count: int) -> str:
    """Fold an image-count placeholder into the persisted turn text.

    The DB `conversations.user_message` column is the ONLY thing history
    replay reads, so recording '[image attached]' here makes past turns show
    that a picture was present with zero image bytes re-sent — the whole
    token-cost goal. The live in-memory user_message stays clean.
    """
    if image_count <= 0:
        return user_message
    label = (
        "[image attached]" if image_count == 1
        else f"[{image_count} images attached]"
    )
    base = user_message.strip()
    return f"{base} {label}" if base else label


async def _describe_images_via_vision(
    images: list[ImageBlock], question: str,
) -> str:
    """Text-describe pasted images through the `vision` profile.

    Used only on the fallback path (chat model is text-only). Caller MUST
    have already confirmed `has_vision_profile(settings)` — get_chat_client
    raises otherwise. Best-effort: a failing description degrades to a marker
    rather than aborting the whole turn.
    """
    client = get_chat_client("vision")
    prompt = (
        "Describe this image in enough detail that a text-only assistant "
        "could answer questions about it: transcribe any visible text, and "
        "note figures, layout, and notable visual content."
    )
    if question.strip():
        prompt += f"\nThe user's question about the image(s): {question.strip()}"
    parts: list[str] = []
    for idx, image in enumerate(images[:_VISION_FALLBACK_MAX_IMAGES], start=1):
        try:
            resp = await client.complete(ChatRequest(
                system="You are a precise image describer.",
                messages=[ChatMessage(
                    role="user",
                    content=[TextBlock(text=prompt), image],
                )],
                max_tokens=_VISION_FALLBACK_MAX_TOKENS,
                temperature=0.0,
            ))
            text = (resp.text or "").strip()
            if not text:
                # A reasoning VLM that hit the token cap before emitting visible
                # text returns empty — log it so this is diagnosable rather than
                # a silent "(description unavailable)".
                log.warning(
                    "vision fallback: empty description for image %d "
                    "(stop_reason=%s, out_tokens=%s) — model may have spent the "
                    "token budget on reasoning; raise vision max_tokens",
                    idx, resp.stop_reason, resp.usage.output_tokens,
                )
        except Exception:
            log.exception("vision fallback: image %d description failed", idx)
            text = ""
        label = f"Image {idx}" if len(images) > 1 else "Image"
        parts.append(f"{label}: {text}" if text else f"{label}: (description unavailable)")
    return "\n\n".join(parts)


def _vision_fallback_user_message(user_message: str, description: str) -> str:
    """Append the auto-description (or a no-vision notice) under a clear
    marker so both plan and execute inherit it as plain text."""
    if description.strip():
        injected = f"{_VISION_FALLBACK_MARKER}\n{description.strip()}"
    else:
        injected = "[image attached, but no vision-capable model is configured]"
    base = user_message.strip()
    return f"{base}\n\n{injected}" if base else injected


async def run_turn(
    *,
    session_id: str,
    user_message: str,
    images: list[ImageBlock] | None = None,
    options: RunOptions | None = None,
    capacity_reserved: bool = False,
) -> AsyncIterator[AgentEvent]:
    """Run one user turn as an event stream.

    Yields AgentEvent frames covering the full plan-execute lifecycle.
    See AgentEvent docstring for event_type semantics.

    `images` (when present) are the CURRENT turn's pasted/dropped images.
    They ride the live user ChatMessage as ImageBlocks (native vision) and
    are never persisted as bytes — only an '[image attached]' placeholder is
    written to history, so later turns never re-pay vision tokens.
    """
    images = list(images or [])
    if not user_message.strip() and not images:
        raise AgentTurnError("user_message is empty")
    options = options or RunOptions()
    settings = get_settings()

    # Decide how pasted images reach the model (settings.chat_vision):
    #   on   → send directly; off → always describe; auto → probe once + cache.
    # On the describe path the raw ImageBlocks are dropped and a vision-profile
    # description is injected as text so both plan and execute inherit it.
    turn_images = images
    turn_user_message = user_message
    if images:
        mode = settings.chat_vision
        if mode == "on":
            send_direct = True
        elif mode == "off":
            send_direct = False
        else:
            send_direct = await _chat_model_supports_vision()
        if not send_direct:
            turn_images = []
            description = ""
            if has_vision_profile(settings):
                description = await _describe_images_via_vision(images, user_message)
            turn_user_message = _vision_fallback_user_message(user_message, description)

    async with session_scope() as db:
        from library.capacity import enforce_chat_concurrency

        stale_seconds = (
            settings.agent_turn_timeout_seconds
            if settings.agent_turn_timeout_seconds > 0
            else 86_400.0
        )
        if not capacity_reserved:
            await enforce_chat_concurrency(
                db,
                limit=settings.chat_concurrency_limit,
                stale_before=datetime.now(timezone.utc) - timedelta(
                    seconds=max(300.0, stale_seconds)
                ),
            )
        last = await session_service.latest_turn_index(db, session_id)
        # Explicit None check — `last or -1` would treat turn_index 0 as
        # falsy and re-issue 0 for the second turn, colliding with the
        # UNIQUE(session_id, turn_index) constraint. (Long-standing bug
        # masked by the previous non-unique index; second turns silently
        # overwrote turn 0 in any read that joined on (session, turn).)
        turn_index = 0 if last is None else last + 1

        conv = await session_service.start_conversation(
            db, session_id=session_id, turn_index=turn_index,
            # Persist a text placeholder (never image bytes) so history
            # replay stays text-only automatically. Count from the original
            # request even on the vision-fallback path.
            user_message=_persisted_user_message(user_message, len(images)),
        )
        # Need session.started_at to freeze the journal slice in the
        # snapshot — see stable_context module docstring.
        session_row = await db.get(SessionRow, session_id)
        if session_row is None:
            raise AgentTurnError(f"session {session_id!r} not found")
        snapshot = await build_stable_snapshot(
            db, session_started_at=session_row.started_at,
        )
        await db.commit()
        conversation_id = conv.id

    # Persist the ORIGINAL pasted images to disk for UI-only re-display when
    # the user revisits this session. This is fully decoupled from the LLM
    # message tape (history replay still reads only the '[image attached]'
    # placeholder above and re-sends zero image bytes). Save the ORIGINAL
    # `images`, NOT the possibly-dropped `turn_images` from the vision
    # fallback path. Best-effort: a save failure must never fail the turn.
    if images:
        try:
            save_turn_attachments(conversation_id, images)
        except Exception:
            log.exception(
                "failed to persist chat attachments for conversation %s",
                conversation_id,
            )

    yield AgentEvent(event_type="conversation", data=conversation_id)

    # Two disjoint prompts (kb-lite-style). Each phase only sees the rules
    # that apply to it, so plan can't be tempted to write a markdown answer
    # under "must always use [^a] footnotes" instructions.
    plan_system = render_phase_system_prompt(phase="plan")
    execute_system = render_phase_system_prompt(phase="execute")
    snapshot_messages = build_snapshot_messages(snapshot)
    plan_history = await build_plan_history_messages(
        session_id, current_conversation_id=conversation_id,
    )
    chat = get_chat_client("chat")

    yield AgentEvent(event_type="planning")
    plan_text = await _run_plan_phase(
        chat=chat,
        system_prompt=plan_system,
        prefix_messages=snapshot_messages + plan_history,
        user_message=turn_user_message,
        images=turn_images,
        conversation_id=conversation_id,
        mode=options.mode,
    )
    session_name = _extract_session_name(plan_text)
    if session_name:
        await _store_session_name(session_id, session_name)
    plan_without_session = _strip_session_name_line(plan_text)
    plan_for_execute = _strip_budget_line(plan_without_session)
    planner_no_plan_answer = _extract_no_plan_answer(plan_for_execute)
    no_plan_repaired = False
    if (
        planner_no_plan_answer is not None
        and _requires_library_tools(turn_user_message)
    ):
        no_plan_repaired = True
        plan_without_session = _no_plan_repair_plan(
            turn_user_message,
            mode=options.mode,
        )
        plan_for_execute = _strip_budget_line(plan_without_session)
        await _record_no_plan_repair(
            conversation_id,
            repaired_plan_text=plan_without_session,
            raw_no_plan_answer=planner_no_plan_answer,
        )
    budget_state = _budget_state_for_plan(
        mode=options.mode,
        plan_text=plan_without_session,
    )
    public_plan_text = _public_plan_text(plan_for_execute)
    yield AgentEvent(
        event_type="plan",
        data=_plan_event_payload(public_plan_text, budget_state),
    )

    outcome = _ExecuteOutcome()
    no_plan_answer = (
        None if no_plan_repaired else _extract_no_plan_answer(plan_for_execute)
    )
    if no_plan_answer is not None:
        # Planner declared the user's turn is trivial — skip execute,
        # still emit one fake "thinking" so the SSE stream shape stays
        # consistent for clients, and an "answer" with the planner's text.
        outcome.answer = no_plan_answer
        yield AgentEvent(
            event_type="answer",
            data=await _rewrite_footnotes_for_display(no_plan_answer),
        )
    else:
        # Resume: replay every prior turn into the executor's message
        # tape so it sees the full session arc, not just the current
        # question. The planner already saw a lighter transcript to
        # resolve terse follow-ups without carrying full tool results.
        settings = get_settings()
        capabilities = getattr(chat, "capabilities", None)
        history_token_budget: int | None = None
        history_tokenizer = "utf8_upper_bound"
        if capabilities is not None:
            history_counter = TokenCounter(capabilities.tokenizer)
            history_tokenizer = capabilities.tokenizer
            fixed_messages = [
                *snapshot_messages,
                ChatMessage(
                    role="user",
                    content=_current_user_content(turn_user_message, turn_images),
                ),
                ChatMessage(
                    role="assistant",
                    content="Plan prepared:\n" + (
                        plan_for_execute or "(no specific plan; answer directly)"
                    ),
                ),
            ]
            fixed_tokens = history_counter.text(execute_system)
            fixed_tokens += history_counter.messages(fixed_messages)
            fixed_tokens += history_counter.text(json.dumps([
                {
                    "name": tool.name,
                    "description": tool.description,
                    "input_schema": tool.input_schema,
                }
                for tool in (
                    all_tool_defs() if capabilities.supports_tools else []
                )
            ], ensure_ascii=False, default=str))
            reserve = max(
                settings.agent_execute_max_tokens,
                settings.conversation_compaction_reserve_tokens,
            )
            history_token_budget = max(
                0,
                int(capabilities.context_window) - reserve - fixed_tokens,
            )
        if history_token_budget is not None and history_token_budget < 256:
            resumed_history = []
        else:
            resumed_history = await build_resumed_messages(
                session_id,
                current_conversation_id=conversation_id,
                compaction_enabled=(
                    settings.conversation_compaction_enabled
                    and history_token_budget is not None
                ),
                token_budget=history_token_budget,
                tokenizer=history_tokenizer,
            )
        async for ev in _run_execute_phase(
            chat=chat,
            system_prompt=execute_system,
            prefix_messages=snapshot_messages,
            plan_text=plan_for_execute,
            user_message=turn_user_message,
            images=turn_images,
            conversation_id=conversation_id,
            session_id=session_id,
            outcome=outcome,
            resumed_history=resumed_history,
            options=options,
            budget_state=budget_state,
        ):
            yield ev
        if not outcome.answer.strip() and outcome.error is None:
            outcome.error = _empty_execute_error()
            outcome.answer = outcome.error
            log.error(
                "conversation %s finished execute without answer or error",
                conversation_id,
            )
            yield AgentEvent(event_type="error", data=outcome.error)

    async with session_scope() as db:
        await session_service.finalize_conversation(
            db,
            conversation_id=conversation_id,
            agent_response=outcome.answer,
        )
        # NO_PLAN turns are trivial by definition (greetings, "thanks",
        # tiny pleasantries the planner answered directly with zero tool
        # calls). Reflecting them produces noisy journal entries that
        # crowd out real investigations and burn one reflect-LLM call per
        # turn for no signal. Skip the enqueue and mark the outcome.
        if no_plan_answer is None and outcome.error is None:
            await enqueue(
                db,
                kind=KIND_REFLECT_TURN,
                payload={"conversation_id": conversation_id},
                dedup_key=f"reflect_turn:{conversation_id}",
            )
        outcome_name = (
            "error" if outcome.error is not None
            else "deferred" if outcome.truncated
            else "applied"
        )
        await record_outcome(
            db,
            task_kind="run_turn",
            object_kind="conversation",
            object_id=conversation_id,
            outcome=outcome_name,
            detail={
                "turn_index": turn_index,
                "session_id": session_id,
                "truncated": outcome.truncated,
                "error": outcome.error,
                "no_plan": no_plan_answer is not None,
                "mode": options.mode,
                "budget": budget_state.payload(),
            },
        )
        conv = await session_service.get_conversation(db, conversation_id)
        cache_summary = summarize_llm_calls([
            call for call in (conv.llm_calls or []) if isinstance(call, dict)
        ])
        usage = TurnUsage(
            input_tokens=conv.total_input_tokens or 0,
            prompt_tokens=cache_summary.prompt_tokens,
            output_tokens=conv.total_output_tokens or 0,
            cache_read_tokens=conv.total_cache_read or 0,
            cache_creation_tokens=cache_summary.cache_creation_tokens,
            cache_eligible_prompt_tokens=cache_summary.eligible_prompt_tokens,
            cache_eligible_read_tokens=cache_summary.eligible_read_tokens,
            cache_eligible_estimated_tokens=cache_summary.eligible_estimated_tokens,
            cache_eligible_requests=cache_summary.eligible_requests,
            prompt_prefix_breaks=cache_summary.prefix_breaks,
            tool_calls=conv.total_tool_calls or 0,
            llm_calls=conv.total_llm_calls or 0,
            duration_ms=conv.total_duration_ms or 0,
            # No pricing table exists, so a stored 0 means "never
            # computed" — surface None instead of a fake $0.
            cost_estimate=conv.total_cost_estimate or None,
        )
        await db.commit()

    yield AgentEvent(
        event_type="done",
        data=json.dumps({
            "session_id": session_id,
            "conversation_id": conversation_id,
            "tokens_in": usage.input_tokens,
            "prompt_tokens": usage.prompt_tokens,
            "tokens_out": usage.output_tokens,
            "cache_read": usage.cache_read_tokens,
            "cache_creation": usage.cache_creation_tokens,
            "cache_eligible_prompt_tokens": usage.cache_eligible_prompt_tokens,
            "cache_eligible_read_tokens": usage.cache_eligible_read_tokens,
            "cache_eligible_estimated_tokens": usage.cache_eligible_estimated_tokens,
            "cache_eligible_requests": usage.cache_eligible_requests,
            "cache_prompt_coverage_ratio": (
                usage.cache_eligible_read_tokens / usage.cache_eligible_prompt_tokens
                if usage.cache_eligible_prompt_tokens > 0
                else None
            ),
            "cache_eligible_hit_ratio": (
                min(
                    1.0,
                    usage.cache_eligible_read_tokens
                    / usage.cache_eligible_estimated_tokens,
                )
                if usage.cache_eligible_estimated_tokens > 0
                else None
            ),
            "cache_eligible_reuse_ratio": (
                min(
                    1.0,
                    usage.cache_eligible_read_tokens
                    / usage.cache_eligible_estimated_tokens,
                )
                if usage.cache_eligible_estimated_tokens > 0
                else None
            ),
            "prompt_prefix_breaks": usage.prompt_prefix_breaks,
            "tool_calls": usage.tool_calls,
            "llm_calls": usage.llm_calls,
            "duration_ms": usage.duration_ms,
            "truncated": outcome.truncated,
            "error": outcome.error,
            "session_name": session_name,
            "mode": options.mode,
            "budget": budget_state.payload(),
        }),
    )


# ---- plan -----------------------------------------------------------------

def _extract_session_name(plan_text: str) -> str | None:
    """Return the planner-supplied session title from the final plan line."""
    if not plan_text:
        return None
    for line in reversed(plan_text.splitlines()):
        text = line.strip()
        if not text:
            continue
        if not text.lower().startswith(SESSION_NAME_PREFIX.lower()):
            return None
        raw = text[len(SESSION_NAME_PREFIX):].strip()
        title = _clean_session_name(raw)
        return title or None
    return None


def _strip_session_name_line(plan_text: str) -> str:
    """Remove the final session-name control line before execute consumes it."""
    if not plan_text:
        return plan_text
    lines = plan_text.splitlines()
    idx = len(lines) - 1
    while idx >= 0 and not lines[idx].strip():
        idx -= 1
    if idx >= 0 and lines[idx].strip().lower().startswith(SESSION_NAME_PREFIX.lower()):
        del lines[idx]
    return "\n".join(lines).strip()


def _extract_budget_tier(plan_text: str) -> BudgetTier | None:
    if not plan_text:
        return None
    for raw in plan_text.splitlines():
        text = raw.strip()
        if not text:
            continue
        if not text.startswith(BUDGET_PREFIX):
            return None
        value = text[len(BUDGET_PREFIX):].strip().casefold()
        if value in BUDGET_TIERS:
            return value  # type: ignore[return-value]
        return None
    return None


def _strip_budget_line(plan_text: str) -> str:
    if not plan_text:
        return plan_text
    lines = plan_text.splitlines()
    if lines and lines[0].strip().startswith(BUDGET_PREFIX):
        return "\n".join(lines[1:]).strip()
    return plan_text.strip()


_NUMBERED_LINE_RE = re.compile(r"^\s*\d+[.)]\s*")


def _public_plan_text(plan_text: str) -> str:
    """Return planner text with numbering stripped for the UI list."""
    if not plan_text:
        return plan_text
    plan_text = _strip_budget_line(plan_text)
    if plan_text.lstrip().startswith(NO_PLAN_PREFIX):
        return plan_text.strip()
    public_lines: list[str] = []
    for raw in plan_text.splitlines():
        line = raw.strip()
        if not line:
            continue
        line = _NUMBERED_LINE_RE.sub("", line).strip()
        if line:
            public_lines.append(line)
    return "\n".join(public_lines).strip()


def _tier_limit(tier: BudgetTier, hard_limit: int) -> int:
    if tier == "quick":
        return min(QUICK_EXECUTE_MAX_TURNS, hard_limit)
    if tier == "standard":
        return min(STANDARD_EXECUTE_MAX_TURNS, hard_limit)
    return hard_limit


def _budget_state_for_plan(*, mode: str, plan_text: str) -> _BudgetState:
    hard_limit = max(3, get_settings().agent_execute_max_turns)
    parsed = _extract_budget_tier(plan_text)
    if mode == "quick":
        tier: BudgetTier = "quick"
        limit = _tier_limit(tier, hard_limit)
        return _BudgetState(
            requested_mode=mode,
            initial_tier=tier,
            current_tier=tier,
            limit=limit,
            hard_limit=limit,
            source="manual",
        )
    if mode == "deep":
        tier = "deep"
        return _BudgetState(
            requested_mode=mode,
            initial_tier=tier,
            current_tier=tier,
            limit=hard_limit,
            hard_limit=hard_limit,
            source="manual",
        )
    tier = parsed or "standard"
    return _BudgetState(
        requested_mode="auto",
        initial_tier=tier,
        current_tier=tier,
        limit=_tier_limit(tier, hard_limit),
        hard_limit=hard_limit,
        source="planner" if parsed else "default",
        max_upgrades=AUTO_MAX_BUDGET_UPGRADES,
    )


def _next_budget_tier(tier: BudgetTier) -> BudgetTier | None:
    idx = BUDGET_TIERS.index(tier)
    if idx + 1 >= len(BUDGET_TIERS):
        return None
    return BUDGET_TIERS[idx + 1]


def _try_upgrade_budget(
    budget: _BudgetState,
    *,
    guard: _CallGuard,
    stats: _DispatchStats,
) -> tuple[bool, int | None]:
    if not budget.auto:
        return False, None
    if budget.upgrades >= budget.max_upgrades:
        return False, None
    if budget.limit >= budget.hard_limit:
        return False, None
    if guard.nudged:
        return False, None
    if stats.successful_new_results <= 0:
        return False, None
    next_tier = _next_budget_tier(budget.current_tier)
    if next_tier is None:
        return False, None
    new_limit = _tier_limit(next_tier, budget.hard_limit)
    if new_limit <= budget.limit:
        return False, None
    previous = budget.limit
    budget.current_tier = next_tier
    budget.limit = new_limit
    budget.upgrades += 1
    return True, previous


def _plan_event_payload(plan_text: str, budget: _BudgetState) -> str:
    payload: dict[str, Any] = {"text": plan_text}
    if not plan_text.lstrip().startswith(NO_PLAN_PREFIX):
        payload["budget"] = budget.payload()
    return json.dumps(payload, ensure_ascii=False)


def _execute_system_prompt_with_budget(
    system_prompt: str,
    *,
    limit: int,
    explicit_finalization: bool = False,
) -> str:
    ending = (
        "then call `finish_research`. The next response will compose the complete "
        "final answer with citations."
        if explicit_finalization
        else "then answer."
    )
    return (
        system_prompt
        + "\n\nRuntime budget: this turn has about "
        f"{limit} execute rounds available. Use tools only until enough "
        f"source evidence is collected, {ending}"
    )


def _clean_session_name(raw: str) -> str:
    title = raw.strip().strip("`\"'")
    title = re.sub(r"\s+", " ", title)
    title = re.sub(r"^\s*[-*#]+\s*", "", title)
    if "entry_id=" in title or re.search(r"\b[0-9a-fA-F]{8}-[0-9a-fA-F-]{27,}\b", title):
        return ""
    return title[:MAX_SESSION_NAME_LEN].rstrip()


async def _store_session_name(session_id: str, session_name: str) -> None:
    try:
        async with session_scope() as db:
            await session_service.update_session_name(
                db, session_id=session_id, name=session_name,
            )
            await db.commit()
    except Exception:
        log.exception("failed to store session name for session %s", session_id)


def _extract_no_plan_answer(plan_text: str) -> str | None:
    """Return the trailing answer if `plan_text` is a NO_PLAN fast-path.

    Tolerates leading whitespace and any minor formatting the model puts
    around the marker. Returns None if this is a normal plan (the common
    path), so the caller falls through to execute.
    """
    if not plan_text:
        return None
    stripped = plan_text.lstrip()
    if not stripped.startswith(NO_PLAN_PREFIX):
        return None
    stripped = _strip_session_name_line(stripped)
    answer = stripped[len(NO_PLAN_PREFIX):].strip()
    # Empty answer body is treated as a non-decision — fall back to execute
    # rather than returning a blank response to the user.
    return answer or None


def _strip_leaked_no_plan(answer: str) -> str:
    """Belt-and-suspenders for a model that mistakenly prefixes its
    execute-phase final answer with the NO_PLAN: control marker. The marker
    is plan-only — never user-visible — so we strip it here regardless of
    what the model emitted."""
    if not answer:
        return answer
    stripped = answer.lstrip()
    if stripped.startswith(NO_PLAN_PREFIX):
        return stripped[len(NO_PLAN_PREFIX):].lstrip()
    return answer


def _prefers_zh(text: str) -> bool:
    return any("\u4e00" <= ch <= "\u9fff" for ch in text or "")


def _empty_answer_fallback(user_message: str) -> str:
    return "（没有生成答案）" if _prefers_zh(user_message) else "(no answer)"


def _empty_execute_error() -> str:
    return (
        "Agent execution failed after planning: the model returned no answer "
        "and no tool calls. Please retry; if it repeats, inspect the provider "
        "response and session resume context."
    )


def _looks_like_future_action(text: str) -> bool:
    """Catch narrow action promises that cannot complete finalizing."""
    return bool(_FUTURE_ACTION_RE.search(text.strip()))


def _finalization_error(user_message: str) -> str:
    if _prefers_zh(user_message):
        return (
            "这次调查已完成取证，但模型未能生成可用的最终答案。"
            "请重试；如果问题持续出现，请检查该会话的模型响应记录。"
        )
    return (
        "The investigation completed its evidence phase, but the model did not "
        "produce a usable final answer. Please retry; if this repeats, inspect "
        "the model response records for this conversation."
    )


def _tool_disabled_fallback(user_message: str) -> str:
    if _prefers_zh(user_message):
        return (
            "模型在工具已禁用后仍尝试调用工具，因此这轮没有生成可靠答案。"
            "请切换到深度模式，或缩小问题范围后重试。"
        )
    return (
        "The model attempted to call a tool after tool use was disabled, so "
        "no reliable final answer was produced. Try Deep mode or narrow the "
        "question."
    )


def _turn_budget_fallback(user_message: str) -> str:
    if _prefers_zh(user_message):
        return "这次调查在生成完整答案前已达到轮次预算。请缩小问题范围，或换一个角度重试。"
    return (
        "This investigation exceeded the turn budget before a complete answer "
        "was produced. Please narrow the question or try another angle."
    )


def _filtered_stop_fallback(user_message: str) -> str:
    if _prefers_zh(user_message):
        return (
            "模型这轮没有返回答案，响应可能被服务商过滤或拒绝。"
            "请调整问题后重试。"
        )
    return (
        "The model ended this turn without an answer; the response was likely "
        "filtered or refused by the provider. Please rephrase and try again."
    )


def _joined_final_answer(parts: list[str], fallback: str = "(no answer)") -> str:
    """Join final-answer fragments from max_tokens continuation calls."""
    text = "".join(p for p in parts if p)
    return _strip_leaked_no_plan(text or fallback)


def _cap_final_answer(answer: str, max_chars: int) -> tuple[str, bool]:
    if max_chars <= 0 or len(answer) <= max_chars:
        return answer, False
    return answer[:max_chars].rstrip(), True


def _fit_provider_messages(
    *,
    chat: Any,
    system_prompt: str,
    messages: list[ChatMessage],
    max_tokens: int,
    tools: list[Any] | None,
    preserved_prefix_count: int = 0,
) -> tuple[list[ChatMessage], dict[str, Any]]:
    """Fit one provider request to the resolved model's hard context limit.

    The stable snapshot prefix is budgeted but never summarized, preserving its
    prompt-cache identity. Conversation history and the growing current-turn
    transcript are compacted as atomic tool exchanges when necessary.
    """
    capabilities = getattr(chat, "capabilities", None)
    if capabilities is None:
        return messages, {
            "conversation_compacted": False,
            "conversation_tokens_before": None,
            "conversation_tokens_after": None,
        }

    counter = TokenCounter(capabilities.tokenizer)
    tool_payload = [
        {
            "name": tool.name,
            "description": tool.description,
            "input_schema": tool.input_schema,
        }
        for tool in (tools or [])
    ]
    fixed_tokens = counter.text(system_prompt)
    fixed_tokens += counter.text(
        json.dumps(tool_payload, ensure_ascii=False, default=str)
    )
    message_budget = int(capabilities.context_window) - max(1, int(max_tokens)) - fixed_tokens
    if message_budget < 128:
        raise ValueError(
            "model context window is too small for the system prompt, tools, "
            "and output reserve"
        )

    prefix_count = max(0, min(int(preserved_prefix_count), len(messages)))
    prefix = messages[:prefix_count]
    dynamic = messages[prefix_count:]
    prefix_tokens = counter.messages(prefix)
    dynamic_budget = message_budget - prefix_tokens
    if dynamic_budget < 128:
        raise ValueError(
            "model context window is too small for the stable prompt prefix "
            "and output reserve"
        )

    tokens_before = counter.messages(messages)
    fitted, checkpoint = fit_messages_to_token_budget(
        dynamic,
        token_budget=dynamic_budget,
        counter=counter,
    )
    result = [*prefix, *fitted]
    tokens_after = counter.messages(result)
    return result, {
        "conversation_compacted": checkpoint is not None,
        "conversation_tokens_before": tokens_before,
        "conversation_tokens_after": tokens_after,
        "conversation_token_budget": message_budget,
        "conversation_tokenizer": counter.tokenizer,
        "conversation_exact_token_count": counter.exact,
        "conversation_checkpoint_summary_tokens": (
            checkpoint.summary_tokens if checkpoint is not None else None
        ),
        "conversation_checkpoint_split_turn": (
            checkpoint.split_turn if checkpoint is not None else None
        ),
    }


async def _run_plan_phase(
    *,
    chat,
    system_prompt: str,
    user_message: str,
    conversation_id: str,
    mode: str = "deep",
    prefix_messages: list[ChatMessage] | None = None,
    images: list[ImageBlock] | None = None,
) -> str:
    settings = get_settings()
    messages = list(prefix_messages or []) + [
        ChatMessage(role="user", content=_current_user_content(user_message, images)),
    ]
    request_messages, compaction_metrics = _fit_provider_messages(
        chat=chat,
        system_prompt=system_prompt,
        messages=messages,
        max_tokens=settings.agent_plan_max_tokens,
        tools=None,
        # The first message is the stable snapshot. Any following planner
        # history is conversation context and may be compacted.
        preserved_prefix_count=1 if prefix_messages else 0,
    )
    started = time.monotonic()
    resp = await chat.complete(ChatRequest(
        system=system_prompt,
        messages=request_messages,
        max_tokens=settings.agent_plan_max_tokens,
        tools=None,            # Plan phase: zero tools (design §10.2).
        json_schema=None,
        cache_breakpoints=[0] if prefix_messages else [],
        temperature=0.3,
    ))
    duration_ms = int((time.monotonic() - started) * 1000)
    plan_text = resp.text or ""
    stored_plan_text = _strip_session_name_line(plan_text)
    async with session_scope() as db:
        await session_service.append_llm_call(
            db,
            conversation_id=conversation_id,
            phase="plan",
            model=getattr(chat, "model", "?"),
            input_tokens=resp.usage.input_tokens,
            prompt_tokens=resp.usage.prompt_tokens,
            output_tokens=resp.usage.output_tokens,
            cache_read_tokens=resp.usage.cache_read_tokens,
            cache_creation_tokens=resp.usage.cache_creation_tokens,
            duration_ms=duration_ms,
            extra={
                "plan_text": stored_plan_text,
                "mode": mode,
                **compaction_metrics,
            },
        )
        await db.commit()
    return plan_text


# ---- live-render footnote rewrite ----------------------------------------

# Agent emits citation defs as:
#     [^a]: entry_id=<id>, quote="<verbatim excerpt>" - reason
#     [^a]: entry_id=<id>, page=<n> - reason
#
# The GUI deep-links via `?q=<urlencoded>` for quote-bearing footnotes (a
# DOM text search highlights the match) and `?page=<n>` for PDFs (the
# browser PDF viewer scrolls). Legacy fields (`lines=`, `section_id=`,
# descriptive `lines=...`) are still tolerated by the regex so historical
# turns don't crash on replay/export, but they don't produce any query
# string — the link opens the file without a jump.
#
# `<id>` accepts a full uuid or a hex-only short prefix (>= 8 chars).
# Backticks around the id / page / quote are tolerated. Quote bodies use
# `\"` and `\\` for embedded `"` and `\`.
#
# Defence in depth against LLM-emitted variants the prompt forbids but can
# still slip through: after `entry_id=<id>`, all extra key/value parameters
# are parsed leniently. Known fields (`quote`, `page`, `section_id`,
# `reason`) are extracted; unknown fields are ignored so they cannot leak as
# raw footnote definitions in the UI.
_LIVE_FOOTNOTE_RE = CITATION_FOOTNOTE_RE


def _parse_live_footnote(match: re.Match[str]) -> CitationFootnote:
    return parse_citation_footnote_match(match)


def _unescape_quote(s: str) -> str:
    return unescape_citation_quote(s)


# Kinds whose FileViewer body is DOM-rendered text — the in-page text
# search behind `?q=<text>` actually scrolls + highlights on these. PDFs
# render in an `<iframe>` that only honours `#page=N`, so they're handled
# separately by mime/extension below.
_TEXT_SEARCHABLE_KINDS = frozenset({"text", "code", "log", "docx"})
_TEXT_SEARCHABLE_EXTS = frozenset({
    "txt", "md", "markdown", "rst", "log", "csv", "tsv",
    "json", "yaml", "yml", "toml", "ini", "conf", "env",
    "sql", "html", "css", "scss", "ts", "tsx", "js", "jsx",
    "py", "rb", "go", "rs", "java", "c", "h", "cpp", "hpp",
    "sh", "bash", "zsh", "ps1", "docx", "pptx", "pptm",
})
_PRESENTATION_EXTS = frozenset({"pptx", "pptm"})
_SPREADSHEET_EXTS = frozenset({"xlsx", "xlsm"})
_EPUB_MIMES = frozenset({"application/epub+zip"})


@dataclass(frozen=True, slots=True)
class _SpreadsheetLocator:
    sheet: str
    cell: str | None = None
    row: int | None = None


@dataclass(frozen=True, slots=True)
class _SpreadsheetTextUnit:
    sheet: str
    text: str
    cell: str | None = None
    row: int | None = None


def _is_pdf_file(file: Any) -> bool:
    if file is None:
        return False
    mime = (getattr(file, "mime_type", None) or "").lower()
    ext = (getattr(file, "original_ext", None) or "").lower().lstrip(".")
    return mime == "application/pdf" or ext == "pdf"


def _is_docx_file(file: Any) -> bool:
    if file is None:
        return False
    mime = (getattr(file, "mime_type", None) or "").lower()
    ext = (getattr(file, "original_ext", None) or "").lower().lstrip(".")
    return ext == "docx" or "wordprocessingml.document" in mime


def _is_presentation_file(file: Any) -> bool:
    if file is None:
        return False
    mime = (getattr(file, "mime_type", None) or "").lower()
    ext = (getattr(file, "original_ext", None) or "").lower().lstrip(".")
    return ext in _PRESENTATION_EXTS or "presentationml.presentation" in mime


def _is_spreadsheet_file(file: Any) -> bool:
    if file is None:
        return False
    mime = (getattr(file, "mime_type", None) or "").lower()
    ext = (getattr(file, "original_ext", None) or "").lower().lstrip(".")
    return ext in _SPREADSHEET_EXTS or "spreadsheetml.sheet" in mime


def _citation_query_string(params: list[tuple[str, object | None]]) -> str:
    query: list[tuple[str, str]] = []
    for key, value in params:
        if value is None:
            continue
        if isinstance(value, int) and value <= 0:
            continue
        text = str(value)
        if text:
            query.append((key, text))
    return f"?{urllib.parse.urlencode(query)}" if query else ""


def _quote_query_value(quote: str | None) -> str | None:
    if not quote:
        return None
    text = _unescape_quote(quote)
    return text if text.strip() else None


def _pick_query_string(
    file: Any,
    quote: str | None,
    page: str | None,
    *,
    located_pdf_page: int | None = None,
    located_docx_block: int | None = None,
    located_pptx_slide: int | None = None,
    located_xlsx_cell: _SpreadsheetLocator | None = None,
) -> str:
    """Build the complete viewer locator supported by the file's real type."""
    if file is None:
        return ""
    query_quote = _quote_query_value(quote)
    if _is_pdf_file(file):
        target_page = located_pdf_page or first_page_number(page)
        if target_page is None:
            return ""
        return _citation_query_string([("page", target_page), ("q", query_quote)])
    mime = (getattr(file, "mime_type", None) or "").lower()
    ext = (getattr(file, "original_ext", None) or "").lower().lstrip(".")
    kind = (getattr(file, "kind", None) or "").lower()
    if query_quote and (ext == "epub" or mime in _EPUB_MIMES):
        return _citation_query_string([("q", query_quote)])
    if _is_docx_file(file):
        return _citation_query_string(
            [("block", located_docx_block), ("q", query_quote)]
            if located_docx_block
            else [("q", query_quote)]
        )
    if _is_presentation_file(file):
        target_slide = located_pptx_slide or first_page_number(page)
        return _citation_query_string([("page", target_slide), ("q", query_quote)])
    if _is_spreadsheet_file(file):
        if located_xlsx_cell is not None:
            if located_xlsx_cell.cell:
                return _citation_query_string([
                    ("sheet", located_xlsx_cell.sheet),
                    ("cell", located_xlsx_cell.cell),
                    ("q", query_quote),
                ])
            return _citation_query_string([
                ("sheet", located_xlsx_cell.sheet),
                ("row", located_xlsx_cell.row),
                ("q", query_quote),
            ])
        return _citation_query_string([("q", query_quote)])
    is_text_searchable = (
        kind in _TEXT_SEARCHABLE_KINDS
        or ext in _TEXT_SEARCHABLE_EXTS
        or mime.startswith("text/")
    )
    if is_text_searchable and query_quote:
        return _citation_query_string([("q", query_quote)])
    return ""


async def _locate_pdf_quote_page(
    file: Any,
    quote: str,
    *,
    pages_cache: dict[str, PdfTextRange] | None = None,
) -> int | None:
    storage_key = getattr(file, "storage_key", None)
    if not storage_key or not quote.strip():
        return None
    try:
        cache_key = str(storage_key)
        if pages_cache is not None and cache_key in pages_cache:
            doc = pages_cache[cache_key]
        else:
            from library.storage import get_storage

            storage = get_storage()
            doc = await get_pdf_text_for_file(storage, file)
            if pages_cache is not None:
                pages_cache[cache_key] = doc
    except Exception:
        log.exception("footnote rewrite: PDF quote locator failed")
        return None

    return locate_quote_page(doc, quote)


async def _resolve_pdf_page_locator(file: Any, page: str | None) -> int | None:
    first = first_page_number(page)
    if first is None:
        return None
    try:
        from library.storage import get_storage

        labels = await get_pdf_page_labels_for_file(get_storage(), file)
        return resolve_page_label(labels, first) or first
    except Exception:
        log.exception("footnote rewrite: PDF page-label lookup failed")
        return first


def _locator_cache_key(file: Any) -> str | None:
    for attr in ("sha256", "id", "storage_key"):
        value = getattr(file, attr, None)
        if value:
            return f"{attr}:{value}"
    return None


async def _read_file_body_for_locator(file: Any) -> bytes:
    storage_key = getattr(file, "storage_key", None)
    if not storage_key:
        raise ValueError("file has no storage_key")
    from library.pipelines.pdf_text import read_storage_bytes
    from library.storage import get_storage

    return await read_storage_bytes(get_storage(), str(storage_key))


async def _locate_docx_quote_block(
    file: Any,
    quote: str,
    *,
    blocks_cache: dict[str, list[str]] | None = None,
) -> int | None:
    if not quote.strip():
        return None
    try:
        cache_key = _locator_cache_key(file)
        if blocks_cache is not None and cache_key and cache_key in blocks_cache:
            blocks = blocks_cache[cache_key]
        else:
            body = await _read_file_body_for_locator(file)
            from library.pipelines.docx import DocxPipeline

            blocks = await asyncio.to_thread(DocxPipeline._parse_paragraphs_from_bytes, body)
            if blocks_cache is not None and cache_key:
                blocks_cache[cache_key] = blocks
    except Exception:
        log.exception("footnote rewrite: DOCX quote locator failed")
        return None

    needle = _unescape_quote(quote)
    for index, block_text in enumerate(blocks, start=1):
        if quote_matches_source_text(block_text, needle):
            return index
    return None


async def _locate_pptx_quote_slide(
    file: Any,
    quote: str,
    *,
    slides_cache: dict[str, list[str]] | None = None,
) -> int | None:
    if not quote.strip():
        return None
    try:
        cache_key = _locator_cache_key(file)
        if slides_cache is not None and cache_key and cache_key in slides_cache:
            slides = slides_cache[cache_key]
        else:
            body = await _read_file_body_for_locator(file)
            from library.pipelines.pptx import MAX_PPTX_SLIDES, PptxPipeline

            slides, _coverage = await asyncio.to_thread(
                PptxPipeline._render_from_bytes_with_coverage,
                body,
                max_slides=MAX_PPTX_SLIDES,
            )
            if slides_cache is not None and cache_key:
                slides_cache[cache_key] = slides
    except Exception:
        log.exception("footnote rewrite: PPTX quote locator failed")
        return None

    needle = _unescape_quote(quote)
    for index, slide_text in enumerate(slides, start=1):
        if quote_matches_source_text(slide_text, needle):
            return index
    return None


async def _locate_xlsx_quote_cell(
    file: Any,
    quote: str,
    *,
    units_cache: dict[str, list[_SpreadsheetTextUnit]] | None = None,
) -> _SpreadsheetLocator | None:
    if not quote.strip():
        return None
    try:
        cache_key = _locator_cache_key(file)
        if units_cache is not None and cache_key and cache_key in units_cache:
            units = units_cache[cache_key]
        else:
            body = await _read_file_body_for_locator(file)
            units = await asyncio.to_thread(_extract_spreadsheet_text_units, body)
            if units_cache is not None and cache_key:
                units_cache[cache_key] = units
    except Exception:
        log.exception("footnote rewrite: XLSX quote locator failed")
        return None

    needle = _unescape_quote(quote)
    for unit in units:
        if unit.cell and quote_matches_source_text(unit.text, needle):
            return _SpreadsheetLocator(sheet=unit.sheet, cell=unit.cell, row=unit.row)
    for unit in units:
        if unit.cell is None and quote_matches_source_text(unit.text, needle):
            return _SpreadsheetLocator(sheet=unit.sheet, row=unit.row)
    return None


def _extract_spreadsheet_text_units(body: bytes) -> list[_SpreadsheetTextUnit]:
    from io import BytesIO

    import openpyxl

    workbook = openpyxl.load_workbook(BytesIO(body), data_only=True, read_only=True)
    try:
        units: list[_SpreadsheetTextUnit] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                row_values: list[str] = []
                row_no: int | None = None
                row_has_value = False
                for cell in row:
                    if row_no is None:
                        row_no = getattr(cell, "row", None) or row_index
                    value = getattr(cell, "value", None)
                    text = "" if value is None else str(value)
                    row_values.append(text)
                    if value is None:
                        continue
                    row_has_value = True
                    coordinate = getattr(cell, "coordinate", None)
                    if coordinate:
                        units.append(_SpreadsheetTextUnit(
                            sheet=sheet_name,
                            cell=str(coordinate),
                            row=row_no,
                            text=text,
                        ))
                if row_has_value:
                    units.append(_SpreadsheetTextUnit(
                        sheet=sheet_name,
                        row=row_no or row_index,
                        text=" ".join(value for value in row_values if value.strip()),
                    ))
        return units
    finally:
        workbook.close()


def _footnote_detail(reason: str | None, quote: str | None) -> str | None:
    parts: list[str] = []
    if quote:
        text = _unescape_quote(quote).strip()
        if text:
            parts.append(f'"{text}"')
    if reason:
        parts.append(reason)
    return " — ".join(parts) if parts else None


_MD_LABEL_ESCAPE_RE = re.compile(r"([\\\[\]()`])")


def _escape_markdown_label(name: str) -> str:
    """Neutralize markdown-significant chars in a display name so a crafted
    filename like `report](https://evil.example)` can't inject a live link
    (or break the citation) when interpolated into `[name](entry:...)`.
    Newlines/whitespace runs collapse to a single space first."""
    return _MD_LABEL_ESCAPE_RE.sub(r"\\\1", " ".join(name.split()))


async def _rewrite_footnotes_for_display(
    answer: str,
    *,
    locate_pdf_quotes: bool = True,
    resolve_pdf_page_labels: bool = True,
) -> str:
    """Resolve `[^a]: entry_id=<uuid>, quote="...", page=N - reason` defs to
    file-type-aware links with exact PDF and Office locators for live rendering.

    The persisted `agent_response` keeps the raw form so downstream exports
    still parse. Missing/ambiguous ids fall back to `(entry <short> unavailable)`.
    Legacy `lines=`/`section_id=` fields are tolerated but don't produce a
    deep-link query string. Locator selection (page vs quote vs bare) is
    driven by the entry's actual file type — see [[_pick_query_string]].
    Replay callers can disable source reads and PDF page-label lookup when
    latency matters more than reconstructing exact locators.
    """
    if not answer or "entry_id" not in answer:
        return answer
    footnotes = [
        _parse_live_footnote(match)
        for match in _LIVE_FOOTNOTE_RE.finditer(answer)
    ]
    if not footnotes:
        return answer

    raw_ids = list({footnote.entry_id for footnote in footnotes})
    name_by_id: dict[str, str] = {}
    file_by_id: dict[str, Any] = {}
    resolved: dict[str, str] = {}
    try:
        async with session_scope() as db:
            for raw in raw_ids:
                full, err = await entries_repo.resolve_entry_id_prefix(db, raw)
                if err is None:
                    resolved[raw] = full
            if resolved:
                rows = await entries_repo.list_live_with_file_by_ids(
                    db, list(set(resolved.values())),
                )
                name_by_id = {entry.id: entry.display_name for entry, _ in rows}
                file_by_id = {entry.id: file for entry, file in rows}
    except Exception:
        log.exception("footnote rewrite: entry lookup failed; keeping raw form")
        return answer

    located_pdf_pages: dict[int, int] = {}
    located_pdf_quotes: set[int] = set()
    located_docx_blocks: dict[int, int] = {}
    located_pptx_slides: dict[int, int] = {}
    located_xlsx_cells: dict[int, _SpreadsheetLocator] = {}
    pdf_pages_cache: dict[str, PdfTextRange] = {}
    docx_blocks_cache: dict[str, list[str]] = {}
    pptx_slides_cache: dict[str, list[str]] = {}
    xlsx_units_cache: dict[str, list[_SpreadsheetTextUnit]] = {}
    for footnote in footnotes:
        raw_eid = footnote.entry_id
        full_eid = resolved.get(raw_eid, raw_eid)
        file = file_by_id.get(full_eid)
        quote = footnote.quote
        page = footnote.page
        if _is_pdf_file(file):
            located = None
            if locate_pdf_quotes and quote:
                located = await _locate_pdf_quote_page(
                    file, quote, pages_cache=pdf_pages_cache,
                )
                if located is not None:
                    located_pdf_quotes.add(footnote.start)
            if located is None and page and resolve_pdf_page_labels:
                located = await _resolve_pdf_page_locator(file, page)
            if located:
                located_pdf_pages[footnote.start] = located
            continue
        if locate_pdf_quotes and quote:
            if _is_docx_file(file):
                block = await _locate_docx_quote_block(
                    file, quote, blocks_cache=docx_blocks_cache,
                )
                if block:
                    located_docx_blocks[footnote.start] = block
            elif _is_presentation_file(file):
                slide = await _locate_pptx_quote_slide(
                    file, quote, slides_cache=pptx_slides_cache,
                )
                if slide:
                    located_pptx_slides[footnote.start] = slide
            elif _is_spreadsheet_file(file):
                cell = await _locate_xlsx_quote_cell(
                    file, quote, units_cache=xlsx_units_cache,
                )
                if cell is not None:
                    located_xlsx_cells[footnote.start] = cell
    footnote_by_start = {footnote.start: footnote for footnote in footnotes}

    def _replace(m: re.Match[str]) -> str:
        footnote = footnote_by_start.get(m.start()) or _parse_live_footnote(m)
        marker = footnote.marker
        raw_eid = footnote.entry_id
        quote = footnote.quote
        page = footnote.page
        reason = footnote.reason

        full_eid = resolved.get(raw_eid, raw_eid)
        short = full_eid[:8]
        name = name_by_id.get(full_eid)
        if name is None:
            head = f"(entry {short} unavailable)"
        else:
            file = file_by_id.get(full_eid)
            query_quote = quote
            if quote and _is_pdf_file(file) and m.start() not in located_pdf_quotes:
                query_quote = None
            qs = _pick_query_string(
                file,
                query_quote,
                page,
                located_pdf_page=located_pdf_pages.get(m.start()),
                located_docx_block=located_docx_blocks.get(m.start()),
                located_pptx_slide=located_pptx_slides.get(m.start()),
                located_xlsx_cell=located_xlsx_cells.get(m.start()),
            )
            head = f"[{_escape_markdown_label(name)}](entry:{full_eid}{qs})"
        detail = _footnote_detail(reason, quote)
        if detail:
            return f"[^{marker}]: {head} — {detail}"
        return f"[^{marker}]: {head}"

    return _LIVE_FOOTNOTE_RE.sub(_replace, answer)



# ---- execute --------------------------------------------------------------

async def _run_execute_phase(
    *,
    chat,
    system_prompt: str,
    plan_text: str,
    user_message: str,
    conversation_id: str,
    session_id: str,
    outcome: _ExecuteOutcome,
    prefix_messages: list[ChatMessage] | None = None,
    resumed_history: list[ChatMessage] | None = None,
    options: RunOptions | None = None,
    budget_state: _BudgetState | None = None,
    images: list[ImageBlock] | None = None,
) -> AsyncIterator[AgentEvent]:
    """Execute loop as event stream.

    Yields AgentEvent frames: thinking / tool_call / tool_result / answer.
    Truncation status and final-answer text are written into `outcome`
    instead of mixed into the stream — keeps the public event stream
    clean (no internal sentinels) and lets the caller branch on plain
    Python attributes.

    `resumed_history` (when present) is the replayed prior-turns context
    built by `build_resumed_messages` and is prepended ahead of the
    current turn's user message, with a boundary note baked in by the
    builder.
    """
    options = options or RunOptions()
    budget_state = budget_state or _budget_state_for_plan(
        mode=options.mode,
        plan_text="",
    )
    quick_mode = options.mode == "quick"
    capabilities = getattr(chat, "capabilities", None)
    tools_supported = capabilities is None or bool(capabilities.supports_tools)
    tool_defs = all_tool_defs() if tools_supported else []
    explicit_finalization = any(
        getattr(tool_def, "name", None) == "finish_research"
        or (
            isinstance(tool_def, dict)
            and tool_def.get("name") == "finish_research"
        )
        for tool_def in tool_defs
    )
    ctx = ToolContext(
        session_id=session_id,
        conversation_id=conversation_id,
        user_message=user_message,
    )
    guard = _CallGuard()

    messages: list[ChatMessage] = (
        list(prefix_messages or [])
        + list(resumed_history or [])
        + [
            # ImageBlocks attach ONLY to this current-turn user message.
            # Prior turns replayed via resumed_history stay text-only.
            ChatMessage(role="user", content=_current_user_content(user_message, images)),
            ChatMessage(role="assistant", content=(
                "Plan prepared:\n"
                + (plan_text or "(no specific plan; answer directly)")
            )),
        ]
    )

    settings = get_settings()
    max_execute_turns = budget_state.limit
    hard_execute_turns = budget_state.hard_limit
    max_final_continuations = (
        0 if quick_mode else max(0, settings.agent_final_answer_continue_turns)
    )
    max_final_chars = max(0, settings.agent_final_answer_max_chars)
    max_total_turns = hard_execute_turns + max_final_continuations + (
        QUICK_FORCED_ANSWER_RETRIES if quick_mode else 0
    ) + (MAX_FINALIZATION_ATTEMPTS if explicit_finalization else 0) + (
        MALFORMED_TOOL_ARGUMENT_REPAIR_LIMIT
    )

    last_text: str | None = None
    final_parts: list[str] = []
    final_continuations = 0
    continuing_final_answer = False
    quick_forced_answer_retries = 0
    quick_forced_answer_active = False
    budget_upgrade_notice: dict[str, Any] | None = None
    tool_calls_seen = False
    no_tool_repair_used = False
    malformed_tool_argument_repairs = 0
    answer_phase: AnswerPhase = (
        "researching" if explicit_finalization else "finalizing"
    )
    finalization_attempts = 0
    finalization_prompt_added = False
    citation_manifest: list[dict[str, Any]] = []
    prefix_tracker = PromptPrefixTracker()
    execute_system_prompt = _execute_system_prompt_with_budget(
        system_prompt,
        limit=hard_execute_turns,
        explicit_finalization=explicit_finalization,
    )

    def enter_finalizing(
        *,
        manifest: list[dict[str, Any]] | None = None,
    ) -> None:
        nonlocal answer_phase, finalization_prompt_added
        nonlocal citation_manifest
        answer_phase = "finalizing"
        if manifest is not None:
            citation_manifest = list(manifest)
        if not finalization_prompt_added:
            messages.append(ChatMessage(role="user", content=FINALIZE_RESEARCH_NUDGE))
            finalization_prompt_added = True

    def finalizing_issue(answer: str) -> str | None:
        if _looks_like_future_action(answer):
            return "the response describes future work instead of answering the user"
        return None

    def request_finalization_retry(
        issue: str,
        *,
        response_text: str | None = None,
    ) -> bool:
        nonlocal continuing_final_answer, final_continuations, last_text
        if finalization_attempts >= MAX_FINALIZATION_ATTEMPTS:
            return False
        if response_text:
            messages.append(ChatMessage(role="assistant", content=response_text))
        messages.append(ChatMessage(
            role="user",
            content=FINALIZATION_RETRY_NUDGE.format(issue=issue),
        ))
        continuing_final_answer = False
        final_continuations = 0
        final_parts.clear()
        last_text = None
        return True

    for turn in range(max_total_turns):
        max_execute_turns = budget_state.limit
        effective_execute_turns = max_execute_turns + malformed_tool_argument_repairs
        if (
            explicit_finalization
            and answer_phase == "researching"
            and turn >= effective_execute_turns - 1
        ):
            enter_finalizing()
        if (
            turn >= effective_execute_turns
            and not continuing_final_answer
            and not quick_forced_answer_active
            and not (
                explicit_finalization
                and answer_phase == "finalizing"
                and finalization_attempts < MAX_FINALIZATION_ATTEMPTS
            )
        ):
            break
        auto_budget_final_round = (
            budget_state.auto
            and not continuing_final_answer
            and turn >= effective_execute_turns - 1
        )
        force_final_answer = (
            (explicit_finalization and answer_phase == "finalizing")
            or (
                not explicit_finalization
                and (quick_mode or auto_budget_final_round)
                and not continuing_final_answer
                and (
                    turn >= effective_execute_turns - 1
                    or quick_forced_answer_active
                )
            )
        )

        budget_tail = (
            None
            if continuing_final_answer
            else _budget_tail(
                turn=turn,
                limit=effective_execute_turns,
                mode=options.mode,
                force_final_answer=force_final_answer,
            )
        )
        if budget_tail:
            messages.append(ChatMessage(role="user", content=budget_tail))
        tools_disabled = (
            continuing_final_answer or force_final_answer or not tools_supported
        )
        request_tools = tool_defs if tools_supported else None
        loop_messages, compaction_metrics = _fit_provider_messages(
            chat=chat,
            system_prompt=execute_system_prompt,
            messages=messages,
            max_tokens=settings.agent_execute_max_tokens,
            tools=request_tools,
            preserved_prefix_count=len(prefix_messages or []),
        )
        if compaction_metrics["conversation_compacted"]:
            # Adopt the checkpoint as this turn's in-memory provider view.
            # Later rounds append to that exact view instead of regenerating a
            # different hidden prefix from the full pre-compaction tape. The
            # persisted conversation remains lossless.
            messages[:] = loop_messages

        thinking_payload = {
            "round": max_execute_turns
            if quick_forced_answer_active else turn + 1,
            "limit": effective_execute_turns,
            "hard_limit": hard_execute_turns,
            "final_continuation": continuing_final_answer,
            "mode": options.mode,
            "budget_tier": budget_state.current_tier,
            "budget_initial_tier": budget_state.initial_tier,
            "budget_upgrades": budget_state.upgrades,
            "force_final_answer": force_final_answer,
            "forced_answer_retry": quick_forced_answer_active,
            "answer_phase": answer_phase,
            "finalization_attempt": (
                finalization_attempts + 1
                if explicit_finalization
                and answer_phase == "finalizing"
                and not continuing_final_answer
                else None
            ),
        }
        if budget_upgrade_notice is not None:
            thinking_payload["budget_upgraded"] = True
            thinking_payload["previous_limit"] = budget_upgrade_notice.get(
                "previous_limit"
            )
            budget_upgrade_notice = None
        yield AgentEvent(
            event_type="thinking",
            data=json.dumps(thinking_payload, ensure_ascii=False),
        )

        if (
            explicit_finalization
            and answer_phase == "finalizing"
            and not continuing_final_answer
        ):
            finalization_attempts += 1

        started = time.monotonic()
        resp = await chat.complete(ChatRequest(
            system=execute_system_prompt,
            messages=loop_messages,
            max_tokens=settings.agent_execute_max_tokens,
            tools=request_tools,
            # Keep provider cache-affecting request parameters stable across
            # execute rounds. Finalization calls are rejected below and never
            # reach the tool dispatcher.
            tool_choice="auto" if tools_supported else "none",
            json_schema=None,
            cache_breakpoints=[0] if prefix_messages else [],
            temperature=0.3,
        ))
        duration_ms = int((time.monotonic() - started) * 1000)
        prompt_observation = prefix_tracker.observe(
            system=execute_system_prompt,
            tools=request_tools,
            messages=loop_messages,
            prompt_tokens=resp.usage.prompt_tokens,
            allow_epoch_break=bool(compaction_metrics["conversation_compacted"]),
            break_reason=(
                "conversation_compaction"
                if compaction_metrics["conversation_compacted"]
                else None
            ),
        )

        async with session_scope() as db:
            await session_service.append_llm_call(
                db,
                conversation_id=conversation_id,
                phase="execute",
                model=getattr(chat, "model", "?"),
                input_tokens=resp.usage.input_tokens,
                prompt_tokens=resp.usage.prompt_tokens,
                output_tokens=resp.usage.output_tokens,
                cache_read_tokens=resp.usage.cache_read_tokens,
                cache_creation_tokens=resp.usage.cache_creation_tokens,
                duration_ms=duration_ms,
                extra={
                    "execute_turn": turn,
                    "mode": options.mode,
                    "budget_tier": budget_state.current_tier,
                    "budget_limit": max_execute_turns,
                    "budget_upgrades": budget_state.upgrades,
                    "stop_reason": resp.stop_reason,
                    "final_continuation": continuing_final_answer,
                    "final_continuation_index": final_continuations
                    if continuing_final_answer else None,
                    "tools_disabled": tools_disabled,
                    "tools_supported": tools_supported,
                    "answer_phase": answer_phase,
                    "finalization_attempt": (
                        finalization_attempts
                        if explicit_finalization and answer_phase == "finalizing"
                        else None
                    ),
                    **compaction_metrics,
                    **prompt_observation.payload(),
                },
            )
            await db.commit()

        if resp.tool_calls and tools_disabled:
            log.warning(
                "conversation %s got tool calls while tools disabled in mode=%s",
                conversation_id,
                options.mode,
            )
            if explicit_finalization and answer_phase == "finalizing":
                issue = "the finalizing response attempted to call a tool"
                if request_finalization_retry(issue, response_text=resp.text):
                    continue
                outcome.error = _finalization_error(user_message)
                outcome.answer = outcome.error
                yield AgentEvent(event_type="error", data=outcome.error)
                return
            if resp.text:
                answer = _strip_leaked_no_plan(resp.text)
                outcome.answer = answer
                yield AgentEvent(
                    event_type="answer",
                    data=await _rewrite_footnotes_for_display(answer),
                )
                return
            if (
                quick_mode
                and quick_forced_answer_retries < QUICK_FORCED_ANSWER_RETRIES
            ):
                quick_forced_answer_retries += 1
                quick_forced_answer_active = True
                messages.append(ChatMessage(
                    role="user",
                    content=QUICK_FORCED_ANSWER_NUDGE,
                ))
                continue
            outcome.truncated = True
            answer = _tool_disabled_fallback(user_message)
            outcome.answer = answer
            yield AgentEvent(event_type="answer", data=answer)
            return

        if resp.tool_calls and not tools_disabled:
            tool_calls_seen = True
            has_malformed_tool_arguments = any(
                tc.parse_error is not None for tc in resp.tool_calls
            )
            assistant_blocks: list = []
            if resp.text:
                assistant_blocks.append(TextBlock(text=resp.text))
            for tc in resp.tool_calls:
                assistant_blocks.append(ToolUseBlock(
                    id=tc.id, name=tc.name, arguments=tc.arguments,
                ))
            messages.append(ChatMessage(role="assistant", content=assistant_blocks))

            tool_result_blocks: list[ToolResultBlock] = []
            dispatch_stats = _DispatchStats()
            async for ev in _dispatch_tool_calls(
                tool_calls=resp.tool_calls,
                ctx=ctx,
                conversation_id=conversation_id,
                result_blocks=tool_result_blocks,
                guard=guard,
                stats=dispatch_stats,
                turn=turn,
            ):
                yield ev
            messages.append(ChatMessage(role="tool", content=tool_result_blocks))
            last_text = resp.text or last_text
            if dispatch_stats.finish_research_requested:
                enter_finalizing(
                    manifest=dispatch_stats.citation_manifest,
                )
                continue
            if (
                has_malformed_tool_arguments
                and malformed_tool_argument_repairs
                < MALFORMED_TOOL_ARGUMENT_REPAIR_LIMIT
            ):
                malformed_tool_argument_repairs += 1
                messages.append(ChatMessage(
                    role="user",
                    content=MALFORMED_TOOL_ARGUMENT_NUDGE,
                ))
                continue
            if turn + 1 >= effective_execute_turns - 1:
                upgraded, previous_limit = _try_upgrade_budget(
                    budget_state,
                    guard=guard,
                    stats=dispatch_stats,
                )
                if upgraded:
                    log.info(
                        "conversation %s auto budget upgraded %s -> %s rounds",
                        conversation_id,
                        previous_limit,
                        budget_state.limit,
                    )
                    budget_upgrade_notice = {
                        "previous_limit": previous_limit,
                    }
            continue

        if explicit_finalization and answer_phase == "researching":
            if resp.text:
                messages.append(ChatMessage(role="assistant", content=resp.text))
            messages.append(ChatMessage(role="user", content=PREMATURE_NO_TOOL_NUDGE))
            no_tool_repair_used = True
            last_text = None
            continue

        if resp.text:
            last_text = resp.text
        if continuing_final_answer or final_parts:
            if resp.text:
                final_parts.append(resp.text)
            answer = _joined_final_answer(
                final_parts,
                last_text or _empty_answer_fallback(user_message),
            )
            answer, capped = _cap_final_answer(answer, max_final_chars)
            if capped:
                log.warning(
                    "conversation %s final answer hit max char cap=%d",
                    conversation_id,
                    max_final_chars,
                )
                if explicit_finalization:
                    issue = (
                        "the answer exceeded the final character limit; return "
                        "a shorter complete answer body"
                    )
                    if request_finalization_retry(issue, response_text=resp.text):
                        continue
                    outcome.error = _finalization_error(user_message)
                    outcome.answer = outcome.error
                    yield AgentEvent(event_type="error", data=outcome.error)
                    return
                outcome.truncated = True
                outcome.answer = answer
                yield AgentEvent(
                    event_type="answer",
                    data=await _rewrite_footnotes_for_display(answer),
                )
                return
            if resp.stop_reason in ("end_turn", "stop_sequence"):
                if explicit_finalization:
                    issue = finalizing_issue(answer)
                    if issue:
                        if request_finalization_retry(
                            issue,
                            response_text=resp.text,
                        ):
                            continue
                        outcome.error = _finalization_error(user_message)
                        outcome.answer = outcome.error
                        yield AgentEvent(event_type="error", data=outcome.error)
                        return
                    answer = attach_citation_manifest(answer, citation_manifest)
                outcome.answer = answer
                yield AgentEvent(
                    event_type="answer",
                    data=await _rewrite_footnotes_for_display(answer),
                )
                return
            if resp.stop_reason == "max_tokens":
                if final_continuations >= max_final_continuations:
                    log.warning(
                        "conversation %s final answer hit continuation limit=%d",
                        conversation_id,
                        max_final_continuations,
                    )
                    if explicit_finalization:
                        issue = (
                            "the final answer remained truncated after its "
                            "continuation limit"
                        )
                        if request_finalization_retry(
                            issue,
                            response_text=resp.text,
                        ):
                            continue
                        outcome.error = _finalization_error(user_message)
                        outcome.answer = outcome.error
                        yield AgentEvent(event_type="error", data=outcome.error)
                        return
                    outcome.truncated = True
                    outcome.answer = answer
                    yield AgentEvent(
                        event_type="answer",
                        data=await _rewrite_footnotes_for_display(answer),
                    )
                    return
                final_continuations += 1
                if resp.text:
                    messages.append(ChatMessage(role="assistant", content=resp.text))
                messages.append(ChatMessage(
                    role="user",
                    content=FINAL_ANSWER_CONTINUE_NUDGE,
                ))
                continuing_final_answer = True
                continue

            log.warning(
                "conversation %s final continuation stopped with %s",
                conversation_id,
                resp.stop_reason,
            )
            if explicit_finalization:
                outcome.error = _finalization_error(user_message)
                outcome.answer = outcome.error
                yield AgentEvent(event_type="error", data=outcome.error)
                return
            outcome.answer = answer
            yield AgentEvent(
                event_type="answer",
                data=await _rewrite_footnotes_for_display(answer),
            )
            return

        if resp.stop_reason in ("end_turn", "stop_sequence"):
            raw_answer = resp.text or last_text or ""
            if not raw_answer.strip():
                if explicit_finalization:
                    issue = "the finalizing response was empty"
                    if request_finalization_retry(issue):
                        continue
                    outcome.error = _finalization_error(user_message)
                    outcome.answer = outcome.error
                    yield AgentEvent(event_type="error", data=outcome.error)
                    return
                outcome.error = _empty_execute_error()
                outcome.answer = outcome.error
                log.error(
                    "conversation %s execute returned empty final response "
                    "(stop_reason=%s)",
                    conversation_id,
                    resp.stop_reason,
                )
                yield AgentEvent(event_type="error", data=outcome.error)
                return
            if (
                not tools_disabled
                and not tool_calls_seen
                and not no_tool_repair_used
                and _requires_library_tools(user_message)
            ):
                no_tool_repair_used = True
                messages.append(ChatMessage(role="assistant", content=raw_answer))
                messages.append(ChatMessage(
                    role="user",
                    content=PREMATURE_NO_TOOL_NUDGE,
                ))
                continue
            answer = _strip_leaked_no_plan(raw_answer)
            if explicit_finalization:
                issue = finalizing_issue(answer)
                if issue:
                    if request_finalization_retry(
                        issue,
                        response_text=resp.text,
                    ):
                        continue
                    outcome.error = _finalization_error(user_message)
                    outcome.answer = outcome.error
                    yield AgentEvent(event_type="error", data=outcome.error)
                    return
                answer = attach_citation_manifest(answer, citation_manifest)
            outcome.answer = answer
            yield AgentEvent(
                event_type="answer",
                data=await _rewrite_footnotes_for_display(answer),
            )
            return
        if resp.stop_reason == "max_tokens":
            final_parts.append(resp.text or last_text or "")
            answer = _joined_final_answer(
                final_parts,
                last_text or _empty_answer_fallback(user_message),
            )
            answer, capped = _cap_final_answer(answer, max_final_chars)
            if capped or max_final_continuations <= 0:
                if capped:
                    log.warning(
                        "conversation %s final answer hit max char cap=%d",
                        conversation_id,
                        max_final_chars,
                    )
                else:
                    log.warning(
                        "conversation %s hit max_tokens with continuation disabled",
                        conversation_id,
                    )
                if explicit_finalization:
                    issue = (
                        "the final answer was truncated; return a shorter complete "
                        "answer body"
                    )
                    if request_finalization_retry(issue, response_text=resp.text):
                        continue
                    outcome.error = _finalization_error(user_message)
                    outcome.answer = outcome.error
                    yield AgentEvent(event_type="error", data=outcome.error)
                    return
                outcome.truncated = True
                outcome.answer = answer
                yield AgentEvent(
                    event_type="answer",
                    data=await _rewrite_footnotes_for_display(answer),
                )
                return
            final_continuations += 1
            if resp.text:
                messages.append(ChatMessage(role="assistant", content=resp.text))
            messages.append(ChatMessage(
                role="user",
                content=FINAL_ANSWER_CONTINUE_NUDGE,
            ))
            continuing_final_answer = True
            continue

        # Terminal non-tool stop reason (e.g. OpenAI content_filter or an
        # Anthropic 'refusal', both mapped to 'other'). Nothing was appended to
        # `messages`, so looping would just re-send an identical request and
        # burn the whole round budget on duplicate filtered/refused outputs.
        # Surface what we have and stop instead of looping.
        surfaced = _strip_leaked_no_plan(resp.text or last_text or "")
        if explicit_finalization:
            outcome.error = _finalization_error(user_message)
            outcome.answer = outcome.error
            yield AgentEvent(event_type="error", data=outcome.error)
            return
        if surfaced.strip():
            outcome.truncated = True
            outcome.answer = surfaced
            yield AgentEvent(
                event_type="answer",
                data=await _rewrite_footnotes_for_display(surfaced),
            )
            return
        log.error(
            "conversation %s execute stopped with stop_reason=%s and no answer",
            conversation_id,
            resp.stop_reason,
        )
        outcome.error = _filtered_stop_fallback(user_message)
        outcome.answer = outcome.error
        yield AgentEvent(event_type="error", data=outcome.error)
        return

    log.warning("conversation %s hit agent_execute_max_turns=%d", conversation_id,
                max_execute_turns)
    if explicit_finalization and answer_phase == "finalizing":
        outcome.error = _finalization_error(user_message)
        outcome.answer = outcome.error
        yield AgentEvent(event_type="error", data=outcome.error)
        return
    fallback = _strip_leaked_no_plan(
        last_text or _turn_budget_fallback(user_message)
    )
    outcome.truncated = True
    outcome.answer = fallback
    yield AgentEvent(
        event_type="answer",
        data=await _rewrite_footnotes_for_display(fallback),
    )


def _budget_tail(
    *,
    turn: int,
    limit: int,
    mode: str = "deep",
    force_final_answer: bool = False,
) -> str | None:
    """Return the budget tail message for execute turn `turn` (0-indexed).

    Always show 'rounds used / left'. Once the run enters the last third of
    `limit`, append a wrap-up nudge so the agent stops gathering and writes
    the answer.
    """
    used = turn  # turns already consumed before this call
    left = limit - used
    base = (
        f"[turn tail] tool rounds used {used} / limit {limit} "
        f"(remaining {left})."
    )
    if force_final_answer and mode != "quick":
        return (
            base
            + " Final budget round: do not call tools. Answer from the "
            "evidence already collected. If evidence is insufficient, state "
            "the gap instead of expanding the search."
        )
    if mode == "quick":
        if used + 1 >= limit:
            return (
                base
                + " Quick mode final execute round: do not call tools. "
                "Do not emit text tool-call markup such as DSML, XML, JSON, "
                "or pseudo function calls. "
                "Answer from the evidence already collected. If evidence is "
                "insufficient, state the gap instead of expanding the search."
            )
        return (
            base
            + " Quick mode: use compact tool calls only for missing evidence; "
            "the final execute round will answer without tools."
        )
    # Nudge once we enter the last third of the budget. For limit=15 this
    # fires from turn 10 onwards (matching the original constant).
    nudge_from = (2 * limit) // 3 + 1
    if used + 1 >= nudge_from:
        base += (
            " You are close to the budget limit. Unless one or two key pieces "
            "of evidence are missing, give the final answer from the material "
            "already collected; do not call more tools."
        )
    return base


def _public_tool_call_id(*, turn: int, tool_index: int) -> str:
    """Replay-stable correlation id that does not expose provider ids."""
    return f"turn-{turn + 1}-tool-{tool_index + 1}"


async def _persist_tool_call(
    *,
    conversation_id: str,
    name: str,
    arguments: Any,
    result: Any,
    error: str | None,
    duration_ms: int,
    tool_call_id: str | None = None,
    tool_index: int | None = None,
    turn: int | None = None,
) -> None:
    """Persist one tool_call row in its own transaction. Used by all four
    dispatch paths (unknown / exception / success / dedup-skipped)."""
    async with session_scope() as db:
        await session_service.append_tool_call(
            db,
            conversation_id=conversation_id,
            name=name,
            arguments=arguments,
            result=result,
            error=error,
            duration_ms=duration_ms,
            tool_call_id=tool_call_id,
            tool_index=tool_index,
            turn=turn,
        )
        await db.commit()


async def _load_prior_tool_calls(
    conversation_id: str,
) -> list[Mapping[str, Any]]:
    """Load completed calls before the current assistant tool-call batch."""
    async with session_scope() as db:
        conversation = await db.get(ConversationRow, conversation_id)
        if conversation is None:
            raise ValueError(f"conversation {conversation_id} missing")
        return [
            call
            for call in list(conversation.tool_calls or [])
            if isinstance(call, Mapping)
        ]


def _persisted_tool_call_failed(call: Mapping[str, Any]) -> bool:
    if call.get("error"):
        return True
    result = call.get("result")
    return isinstance(result, Mapping) and (
        result.get("ok") is False or bool(result.get("error"))
    )


def _finish_research_preflight(
    tool_call: Any,
    prior_tool_calls: Sequence[Mapping[str, Any]],
) -> dict[str, Any] | None:
    if tool_call.name != "finish_research":
        return None
    if str(tool_call.arguments.get("evidence_status") or "") != "sufficient":
        return None
    latest_read = next(
        (
            call
            for call in reversed(prior_tool_calls)
            if str(call.get("name") or "") == "read_files"
        ),
        None,
    )
    if latest_read is None or not _persisted_tool_call_failed(latest_read):
        return None
    return {
        "error": (
            "finish_research(sufficient) was rejected because the latest "
            "read_files call failed; resolve or replace that read before finishing"
        ),
        "retryable": True,
        "guard": "unresolved_read_failure",
    }


async def _dispatch_tool_calls(
    *,
    tool_calls,
    ctx: ToolContext,
    conversation_id: str,
    result_blocks: list[ToolResultBlock],
    guard: _CallGuard,
    stats: _DispatchStats | None = None,
    turn: int = 0,
) -> AsyncIterator[AgentEvent]:
    """Preflight + parallel execution + completion-order drain.

    Async generator yielding AgentEvent (`tool_call`, `tool_result`,
    sometimes `user_artifact`). Two invariants worth pinning:

      - SSE events fire in *completion* order — users see fast tools
        finish first regardless of where they were in the assistant
        message. Each event carries `tool_call_id` so the frontend can
        pair the result back to the right step.
      - `result_blocks` is appended in *source* order so Anthropic's
        tool_use_id ↔ tool_result_id pairing stays valid when this
        message is fed back to the model.

    Guards (append-only — never edit prior history):
      - dedup-prior: (name, args) seen earlier this turn → synthesize a
        ToolResultBlock with the prior result_text, skip handler.
      - dedup-batch: (name, args) appears twice in *this* batch → only
        the first runs; the duplicate waits for the leader's result and
        reuses it. Saves real work when the model fans out the same
        read twice in one assistant message.
      - doom-loop: if the same key crossed DOOM_LOOP_THRESHOLD in the
        last DOOM_LOOP_WINDOW dispatched calls, append a STOP nudge to
        the last ToolResultBlock of *this* tool message.
    """
    n = len(tool_calls)
    public_ids = [
        _public_tool_call_id(turn=turn, tool_index=idx)
        for idx in range(n)
    ]
    placeholders: list[ToolResultBlock | None] = [None] * n
    keys: list[str] = []
    statuses: list[str] = []  # runnable | duplicate | unknown | malformed | preflight_error
    leader_followers: dict[int, list[int]] = {}
    seen_in_batch: dict[str, int] = {}
    nudge_pending = False
    prior_tool_calls = (
        await _load_prior_tool_calls(conversation_id)
        if any(tc.name == "finish_research" for tc in tool_calls)
        else []
    )
    finish_manifests: dict[int, list[dict[str, Any]]] = {}
    preflight_errors: dict[int, dict[str, Any]] = {}

    # ---- preflight: classify in source order, yield tool_call events ----
    for idx, tc in enumerate(tool_calls):
        # Pre-resolve every id referenced in args so the display
        # one-liner can show names instead of raw uuids. One DB round
        # trip per tool call, skipped when no ids of that kind appear.
        eids = tool_display.collect_entry_ids(tc.name, tc.arguments)
        tids = tool_display.collect_tag_ids(tc.name, tc.arguments)
        fids = tool_display.collect_folder_ids(tc.name, tc.arguments)
        cids = tool_display.collect_catalog_ids(tc.name, tc.arguments)
        name_by_id: dict[str, str] = {}
        tag_name_by_id: dict[str, str] = {}
        folder_name_by_id: dict[str, str] = {}
        catalog_name_by_id: dict[str, str] = {}
        if eids or tids or fids or cids:
            try:
                async with session_scope() as _db:
                    if eids:
                        # Agent may pass either a full uuid or a short
                        # hex prefix (>= 8 chars). Resolve each first so
                        # the display layer can label both forms with
                        # the entry's display_name.
                        full_by_raw: dict[str, str] = {}
                        for raw in set(eids):
                            full, err = await entries_repo.resolve_entry_id_prefix(
                                _db, raw,
                            )
                            if err is None:
                                full_by_raw[raw] = full
                        if full_by_raw:
                            rows = await entries_repo.list_live_with_file_by_ids(
                                _db, list(set(full_by_raw.values())),
                            )
                            full_to_name = {
                                entry.id: entry.display_name for entry, _ in rows
                            }
                            # Key the resolver by what the agent actually
                            # passed (raw), so format_tool_call can look
                            # up the same id it sees in `args`.
                            name_by_id = {
                                raw: full_to_name[full]
                                for raw, full in full_by_raw.items()
                                if full in full_to_name
                            }
                    if tids:
                        tag_name_by_id = await tags_repo.name_by_ids(
                            _db, list(set(tids))
                        )
                    if fids:
                        folder_name_by_id = await folders_repo.name_by_ids(
                            _db, list(set(fids))
                        )
                    if cids:
                        catalog_name_by_id = await catalogs_repo.name_by_ids(
                            _db, list(set(cids))
                        )
            except Exception:
                log.exception("tool_call display: name lookup failed")

        display = tool_display.format_tool_call(
            tc.name, tc.arguments,
            resolver=name_by_id.get,
            tag_resolver=tag_name_by_id.get,
            folder_resolver=folder_name_by_id.get,
            catalog_resolver=catalog_name_by_id.get,
        )

        yield AgentEvent(
            event_type="tool_call",
            data=json.dumps({
                "tool_call_id": public_ids[idx],
                "tool_index": idx,
                "turn": turn,
                "name": tc.name,
                "arguments": tc.arguments,
                "display": display,
                "entry_names": name_by_id,
                "tag_names": tag_name_by_id,
                "folder_names": folder_name_by_id,
                "catalog_names": catalog_name_by_id,
            }, ensure_ascii=False),
        )

        key = guard.key(tc.name, tc.arguments)
        keys.append(key)

        if getattr(tc, "parse_error", None) is not None:
            statuses.append("malformed")
            continue

        preflight_error = _finish_research_preflight(tc, prior_tool_calls)
        finish_manifest: list[dict[str, Any]] = []
        if preflight_error is None:
            finish_manifest, preflight_error = prepare_finish_citation_manifest(
                tc,
                prior_tool_calls,
            )
        if tc.name == "finish_research":
            finish_manifests[idx] = finish_manifest
        if preflight_error is not None:
            preflight_errors[idx] = preflight_error
            statuses.append("preflight_error")
            continue

        if guard.should_nudge(key):
            nudge_pending = True
            guard.nudged = True

        if tc.name != "finish_research" and guard.is_duplicate(key):
            statuses.append("dup_prior")
            guard.recent.append(key)
            continue
        if tc.name != "finish_research" and key in seen_in_batch:
            statuses.append("dup_batch")
            leader_followers.setdefault(seen_in_batch[key], []).append(idx)
            guard.recent.append(key)
            continue
        if get_tool(tc.name) is None:
            statuses.append("unknown")
            continue
        seen_in_batch[key] = idx
        statuses.append("runnable")

    # ---- synchronous resolution: dup_prior + unknown ----
    for idx, tc in enumerate(tool_calls):
        s = statuses[idx]
        key = keys[idx]
        if s == "dup_prior":
            prior = guard.seen[key]
            prior_preview = guard.seen_previews.get(key) or "(see prior call)"
            if stats is not None and tc.name == "finish_research":
                stats.finish_research_requested = True
                stats.evidence_status = str(
                    tc.arguments.get("evidence_status") or ""
                ) or None
            placeholders[idx] = ToolResultBlock(
                tool_call_id=tc.id,
                content=(
                    "[runtime guard] duplicate call this turn — reusing "
                    f"prior result.\n{prior}"
                ),
            )
            await _persist_tool_call(
                conversation_id=conversation_id,
                name=tc.name,
                arguments=tc.arguments,
                result={"deduped": True, "preview": prior_preview},
                error=None,
                duration_ms=0,
                tool_call_id=public_ids[idx],
                tool_index=idx,
                turn=turn,
            )
            yield AgentEvent(
                event_type="tool_result",
                data=json.dumps({
                    "tool_call_id": public_ids[idx],
                    "tool_index": idx,
                    "turn": turn,
                    "name": tc.name, "ok": True, "deduped": True,
                    "preview": prior_preview[:TOOL_RESULT_PREVIEW_LEN],
                }, ensure_ascii=False),
            )
        elif s == "unknown":
            err = f"unknown tool: {tc.name}"
            await _persist_tool_call(
                conversation_id=conversation_id,
                name=tc.name, arguments=tc.arguments,
                result=None, error=err, duration_ms=0,
                tool_call_id=public_ids[idx], tool_index=idx, turn=turn,
            )
            placeholders[idx] = ToolResultBlock(
                tool_call_id=tc.id,
                content=f"ERROR: {err}",
                is_error=True,
            )
            guard.remember(key, f"ERROR: {err}")
            yield AgentEvent(
                event_type="tool_result",
                data=json.dumps({
                    "tool_call_id": public_ids[idx],
                    "tool_index": idx,
                    "turn": turn,
                    "name": tc.name, "ok": False, "error": err,
                }, ensure_ascii=False),
            )
        elif s == "malformed":
            err = f"could not parse tool arguments: {getattr(tc, 'parse_error', None)}"
            result = {
                "error": err,
                "retryable": True,
                "correction": (
                    "Retry this tool call with one complete JSON object matching "
                    "the tool schema. Do not wrap arguments in a Markdown code fence."
                ),
            }
            await _persist_tool_call(
                conversation_id=conversation_id,
                name=tc.name,
                arguments=tc.arguments,
                result=result,
                error=err,
                duration_ms=0,
                tool_call_id=public_ids[idx],
                tool_index=idx,
                turn=turn,
            )
            result_text = json.dumps(result, ensure_ascii=False)
            placeholders[idx] = ToolResultBlock(
                tool_call_id=tc.id,
                content=result_text,
                is_error=True,
            )
            guard.remember(key, result_text, preview=err)
            yield AgentEvent(
                event_type="tool_result",
                data=json.dumps({
                    "tool_call_id": public_ids[idx],
                    "tool_index": idx,
                    "turn": turn,
                    "name": tc.name,
                    "ok": False,
                    "error": err,
                    "retryable": True,
                }, ensure_ascii=False),
            )
        elif s == "preflight_error":
            result = preflight_errors[idx]
            err = str(result.get("error") or "finish_research preflight failed")
            await _persist_tool_call(
                conversation_id=conversation_id,
                name=tc.name,
                arguments=tc.arguments,
                result=result,
                error=err,
                duration_ms=0,
                tool_call_id=public_ids[idx],
                tool_index=idx,
                turn=turn,
            )
            result_text = json.dumps(result, ensure_ascii=False)
            placeholders[idx] = ToolResultBlock(
                tool_call_id=tc.id,
                content=result_text,
                is_error=True,
            )
            guard.remember(key, result_text, preview=err)
            yield AgentEvent(
                event_type="tool_result",
                data=json.dumps({
                    "tool_call_id": public_ids[idx],
                    "tool_index": idx,
                    "turn": turn,
                    "name": tc.name,
                    "ok": False,
                    "error": err,
                    "guard": result.get("guard"),
                }, ensure_ascii=False),
            )

    # ---- run a bounded rolling pool (each task owns its own DB session) ----
    # A model can emit a large fan-out in one response. Starting every call at
    # once used to consume one database connection and one result buffer per
    # call. Queue the overflow while preserving completion-order SSE events.
    scheduled: list[ScheduledTool[tuple[int, Any, Any]]] = []
    for idx, tc in enumerate(tool_calls):
        if statuses[idx] == "runnable":
            registration = get_tool(tc.name)
            policy = getattr(registration, "policy", None)
            scheduled.append(ScheduledTool(
                index=len(scheduled),
                cache_key=keys[idx],
                concurrency=getattr(policy, "concurrency", "parallel"),
                value=(idx, registration, tc),
            ))

    pending: deque[tuple[int, int, Any, Any]] = deque(
        (wave_number, *call.value)
        for wave_number, wave in enumerate(schedule_waves(scheduled))
        for call in wave
    )

    tasks: dict[asyncio.Task, int] = {}
    max_parallelism = get_settings().agent_max_parallel_tool_calls
    active_wave: int | None = None
    fatal_failure = False

    def fill_task_slots() -> None:
        nonlocal active_wave
        if fatal_failure or not pending:
            return
        if not tasks:
            active_wave = pending[0][0]
        while (
            pending
            and pending[0][0] == active_wave
            and len(tasks) < max_parallelism
        ):
            _wave, idx, reg, tc = pending.popleft()
            tasks[asyncio.create_task(_run_tool(reg, ctx, tc))] = idx

    fill_task_slots()

    # ---- drain in completion order ----
    try:
        while tasks:
            done, _pending = await asyncio.wait(
                list(tasks.keys()),
                return_when=asyncio.FIRST_COMPLETED,
            )
            for task in sorted(done, key=lambda item: tasks[item]):
                idx = tasks.pop(task)
                tc = tool_calls[idx]
                key = keys[idx]
                duration_ms, result, exc = task.result()
                if exc is not None:
                    fatal_failure = True
                    log.exception("tool %s failed", tc.name, exc_info=exc)
                    err = repr(exc)
                    await _persist_tool_call(
                        conversation_id=conversation_id,
                        name=tc.name, arguments=tc.arguments,
                        result=None, error=err, duration_ms=duration_ms,
                        tool_call_id=public_ids[idx], tool_index=idx, turn=turn,
                    )
                    placeholders[idx] = ToolResultBlock(
                        tool_call_id=tc.id,
                        content=f"ERROR: {err}",
                        is_error=True,
                    )
                    guard.remember(key, f"ERROR: {err}")
                    yield AgentEvent(
                        event_type="tool_result",
                        data=json.dumps({
                            "tool_call_id": public_ids[idx],
                            "tool_index": idx,
                            "turn": turn,
                            "name": tc.name, "ok": False, "error": err,
                            "duration_ms": duration_ms,
                        }, ensure_ascii=False),
                    )
                    # Fan-out failures to batch followers too — they share
                    # the leader's verdict.
                    for fidx in leader_followers.get(idx, ()):
                        ftc = tool_calls[fidx]
                        placeholders[fidx] = ToolResultBlock(
                            tool_call_id=ftc.id,
                            content=(
                                "[runtime guard] duplicate call this batch — "
                                f"leader failed.\nERROR: {err}"
                            ),
                            is_error=True,
                        )
                        await _persist_tool_call(
                            conversation_id=conversation_id,
                            name=ftc.name,
                            arguments=ftc.arguments,
                            result={"deduped": True},
                            error=err,
                            duration_ms=duration_ms,
                            tool_call_id=public_ids[fidx],
                            tool_index=fidx,
                            turn=turn,
                        )
                        yield AgentEvent(
                            event_type="tool_result",
                            data=json.dumps({
                                "tool_call_id": public_ids[fidx],
                                "tool_index": fidx,
                                "turn": turn,
                                "name": ftc.name, "ok": False,
                                "deduped": True, "error": err,
                            }, ensure_ascii=False),
                        )
                    continue

                if tc.name == "finish_research" and isinstance(result, dict):
                    manifest = finish_manifests.get(idx, [])
                    result = dict(result)
                    if manifest:
                        result["citation_manifest"] = manifest
                        result["next"] = (
                            "Write the final answer now without calling more tools. "
                            "Use the citation_manifest markers in the body; the "
                            "validated footnote definitions will be appended "
                            "deterministically."
                        )

                # Side-channel: tools may attach `__user_only__` payload
                # shown to the UI but kept OUT of the model's tool_result
                # content. We persist the full result on the conversation
                # row so /info and replays still show it.
                user_only = None
                if isinstance(result, dict) and "__user_only__" in result:
                    user_only = result.get("__user_only__")
                    result_for_model_source = {
                        k: v for k, v in result.items() if k != "__user_only__"
                    }
                else:
                    result_for_model_source = result
                result_ok = not (
                    isinstance(result, dict)
                    and (result.get("ok") is False or result.get("error"))
                )
                if (
                    stats is not None
                    and tc.name == "finish_research"
                    and result_ok
                ):
                    stats.finish_research_requested = True
                    stats.evidence_status = str(
                        tc.arguments.get("evidence_status") or ""
                    ) or None
                    raw_manifest = result_for_model_source.get("citation_manifest")
                    stats.citation_manifest = (
                        [
                            dict(item)
                            for item in raw_manifest
                            if isinstance(item, Mapping)
                        ]
                        if isinstance(raw_manifest, list)
                        else []
                    )
                compressed_for_model = maybe_compress_tool_result_for_model(
                    tc.name,
                    result_for_model_source,
                    context=ctx.user_message,
                )
                model_payload = (
                    compressed_for_model
                    if compressed_for_model is not None
                    else result_for_model_source
                )
                result_for_model = _copy_jsonish(model_payload)
                if isinstance(result_for_model, (dict, list)):
                    result_text, _trim_marker = _structured_truncate(
                        result_for_model, MAX_TOOL_RESULT_LEN,
                    )
                else:
                    result_text = json.dumps(result_for_model, ensure_ascii=False)
                    if len(result_text) > MAX_TOOL_RESULT_LEN:
                        result_text = result_text[:MAX_TOOL_RESULT_LEN] + "...(truncated)"
                await _persist_tool_call(
                    conversation_id=conversation_id,
                    name=tc.name, arguments=tc.arguments,
                    result=result, error=None,
                    duration_ms=duration_ms,
                    tool_call_id=public_ids[idx], tool_index=idx, turn=turn,
                )
                if user_only is not None:
                    yield AgentEvent(
                        event_type="user_artifact",
                        data=json.dumps({
                            "tool_call_id": public_ids[idx],
                            "tool_index": idx,
                            "turn": turn,
                            "tool": tc.name,
                            "payload": user_only,
                        }, ensure_ascii=False),
                    )
                preview = tool_display.format_tool_result_preview(
                    tc.name, result_for_model_source,
                )
                if len(preview) > TOOL_RESULT_PREVIEW_LEN:
                    preview = preview[:TOOL_RESULT_PREVIEW_LEN] + "..."
                placeholders[idx] = ToolResultBlock(
                    tool_call_id=tc.id, content=result_text,
                )
                guard.remember(key, result_text, preview=preview)
                if stats is not None and tc.name != "finish_research":
                    stats.successful_new_results += 1
                yield AgentEvent(
                    event_type="tool_result",
                    data=json.dumps({
                        "tool_call_id": public_ids[idx],
                        "tool_index": idx,
                        "turn": turn,
                        "name": tc.name, "ok": True, "preview": preview,
                        "duration_ms": duration_ms,
                    }, ensure_ascii=False),
                )
                # Fan-out the leader's result to its batch followers.
                for fidx in leader_followers.get(idx, ()):
                    ftc = tool_calls[fidx]
                    placeholders[fidx] = ToolResultBlock(
                        tool_call_id=ftc.id,
                        content=(
                            "[runtime guard] duplicate call this batch — "
                            f"reusing leader's result.\n{result_text}"
                        ),
                    )
                    await _persist_tool_call(
                        conversation_id=conversation_id,
                        name=ftc.name,
                        arguments=ftc.arguments,
                        result={"deduped": True, "preview": preview},
                        error=None,
                        duration_ms=duration_ms,
                        tool_call_id=public_ids[fidx],
                        tool_index=fidx,
                        turn=turn,
                    )
                    yield AgentEvent(
                        event_type="tool_result",
                        data=json.dumps({
                            "tool_call_id": public_ids[fidx],
                            "tool_index": fidx,
                            "turn": turn,
                            "name": ftc.name, "ok": True,
                            "deduped": True,
                            "preview": preview[:TOOL_RESULT_PREVIEW_LEN],
                        }, ensure_ascii=False),
                    )
            fill_task_slots()
    finally:
        for t in tasks:
            t.cancel()
        if tasks:
            await asyncio.gather(*tasks, return_exceptions=True)

    # A fatal executor failure is a stop signal, not permission to silently
    # drop the model's remaining tool calls. Return explicit error results for
    # every call that was deliberately left unstarted so the assistant/tool
    # exchange remains structurally complete.
    while pending:
        _wave, idx, _reg, tc = pending.popleft()
        key = keys[idx]
        err = "tool execution was not started after an earlier tool failed"
        await _persist_tool_call(
            conversation_id=conversation_id,
            name=tc.name,
            arguments=tc.arguments,
            result=None,
            error=err,
            duration_ms=0,
            tool_call_id=public_ids[idx],
            tool_index=idx,
            turn=turn,
        )
        placeholders[idx] = ToolResultBlock(
            tool_call_id=tc.id,
            content=f"ERROR: {err}",
            is_error=True,
        )
        guard.remember(key, f"ERROR: {err}")
        yield AgentEvent(
            event_type="tool_result",
            data=json.dumps({
                "tool_call_id": public_ids[idx],
                "tool_index": idx,
                "turn": turn,
                "name": tc.name,
                "ok": False,
                "error": err,
                "not_started": True,
            }, ensure_ascii=False),
        )
        for follower_idx in leader_followers.get(idx, ()):
            follower = tool_calls[follower_idx]
            placeholders[follower_idx] = ToolResultBlock(
                tool_call_id=follower.id,
                content=(
                    "[runtime guard] duplicate call this batch — leader was "
                    f"not started.\nERROR: {err}"
                ),
                is_error=True,
            )
            await _persist_tool_call(
                conversation_id=conversation_id,
                name=follower.name,
                arguments=follower.arguments,
                result={"deduped": True, "not_started": True},
                error=err,
                duration_ms=0,
                tool_call_id=public_ids[follower_idx],
                tool_index=follower_idx,
                turn=turn,
            )
            yield AgentEvent(
                event_type="tool_result",
                data=json.dumps({
                    "tool_call_id": public_ids[follower_idx],
                    "tool_index": follower_idx,
                    "turn": turn,
                    "name": follower.name,
                    "ok": False,
                    "deduped": True,
                    "error": err,
                    "not_started": True,
                }, ensure_ascii=False),
            )

    # ---- finalize: source-order result_blocks + doom-loop nudge ----
    for ph in placeholders:
        if ph is not None:
            result_blocks.append(ph)

    if nudge_pending and result_blocks:
        # Decorate the last real tool_result with the STOP nudge. We
        # cannot append a synthetic ToolResultBlock with a fake
        # tool_use_id — Anthropic validates ids against prior tool_use
        # blocks and rejects unknown ones. Appending text to an existing
        # block's `content` keeps the message valid AND append-only at
        # the conversation level (we are decorating a block we just
        # created in this turn — never touching history).
        last = result_blocks[-1]
        result_blocks[-1] = ToolResultBlock(
            tool_call_id=last.tool_call_id,
            content=f"{last.content}\n\n{DOOM_LOOP_NUDGE}",
            is_error=last.is_error,
        )


async def _run_tool(reg, ctx: ToolContext, tc) -> tuple[int, Any, Exception | None]:
    """Execute one tool inside its own session_scope. Returns
    (duration_ms, result, exception). Never raises — failures travel
    back as the third tuple element so the dispatcher loop stays clean.
    """
    started = time.monotonic()
    try:
        policy = getattr(reg, "policy", None)
        timeout_seconds = getattr(policy, "timeout_seconds", 120.0)
        async with asyncio.timeout(timeout_seconds):
            async with session_scope() as db:
                async with tool_execution_lock(
                    db,
                    concurrency=getattr(policy, "concurrency", "parallel"),
                    session_id=ctx.session_id,
                    tool_name=getattr(reg, "name", getattr(tc, "name", "tool")),
                ):
                    result = await reg.handler(db, ctx, tc.arguments)
                    await db.commit()
        return int((time.monotonic() - started) * 1000), result, None
    except Exception as exc:  # noqa: BLE001
        return int((time.monotonic() - started) * 1000), None, exc
