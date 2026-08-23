"""Spreadsheet pipeline (.xlsx / .xlsm via openpyxl).

Renders sheets as a markdown-ish text view: one `# Sheet: <name>` heading
per sheet, then up to N rows. Long sheets are sampled — the first M rows
plus a tail summary — so the LLM gets a sense of structure without
swallowing a 100k-row spreadsheet.

read_segment supports `heading="Sheet: <name>"` to fetch one sheet's
rendered rows, `pattern` for regex search, and the generic
offset/max_chars chunking. For tabular querying use `query_sql`.

XLS (legacy binary) is not supported; users should resave to .xlsx.
"""
from __future__ import annotations

import asyncio
import io
import logging
import re
from typing import Any

from library.pipelines._text_indexer import index_extracted_text
from library.pipelines.base import (
    Pipeline,
    PipelineContext,
    PipelineResult,
    SegmentResult,
)
from library.pipelines.registry import register_pipeline
from library.storage.base import StorageBackend
from library.tasks.usage import measure_stage

log = logging.getLogger(__name__)

# Do not reject XLSX by compressed package size. Ingest is controlled by
# sheet/row sampling below, which is a better proxy for work than media size.
MAX_ROWS_PER_SHEET = 200
MAX_TAIL_PEEK = 20
MAX_CELL_CHARS = 200
DEFAULT_MAX_CHARS = 8000


@register_pipeline(
    mimes=(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroenabled.12",
    ),
    exts=(".xlsx", ".xlsm"),
)
class SpreadsheetPipeline(Pipeline):
    name = "spreadsheet"

    async def run(
        self,
        *,
        ctx: PipelineContext,
        storage: StorageBackend,
    ) -> PipelineResult:
        with measure_stage("extraction"):
            body, coverage = await self._extract_text_with_coverage(
                storage, ctx.storage_key,
            )
        fallback_sections = _sheet_sections(coverage)
        return await index_extracted_text(
            body,
            ctx,
            kind="table",
            coverage=coverage,
            fallback_sections=fallback_sections,
            pipeline=self.name,
            metadata={**coverage, "sections": fallback_sections},
        )

    async def read_segment(
        self,
        *,
        file_row: Any,
        args: dict[str, Any],
        storage: StorageBackend,
    ) -> SegmentResult:
        body = await self._extract_read_text(storage, file_row.storage_key)
        return self._slice(body, args)

    async def read_segment_from_bytes(
        self,
        body: bytes,
        args: dict[str, Any],
        *,
        filename: str | None = None,
    ) -> SegmentResult:
        """Bytes-first variant — used by ArchivePipeline for member peeks."""
        try:
            text = self._render_read_from_bytes(body)
        except Exception as exc:  # noqa: BLE001
            return SegmentResult(error=f"xlsx parse failed: {exc}")
        return self._slice(text, args)

    def _slice(self, body: str, args: dict[str, Any]) -> SegmentResult:
        """Resolve args against the rendered workbook.

        Field priority:
          1. pattern              → regex search over rendered text
          2. heading              → "Sheet: <name>" → that sheet's rows
          3. (default)            → offset..offset+max_chars chunk
        """
        offset = max(0, int(args.get("offset") or 0))
        max_chars = int(args.get("max_chars") or DEFAULT_MAX_CHARS)
        if max_chars <= 0:
            max_chars = DEFAULT_MAX_CHARS

        pattern = (args.get("pattern") or "").strip()
        if pattern:
            scope_body = body
            heading_scope = (args.get("heading") or "").strip()
            if heading_scope:
                slab = _slice_by_heading(body, heading_scope)
                if slab is None:
                    sheet_names = _list_sheet_headings(body)
                    return SegmentResult(
                        error=f"sheet/heading not found: {heading_scope!r}",
                        extras={"available_sheets": sheet_names},
                    )
                scope_body = slab
            return _ss_pattern_search(
                body=scope_body, pattern=pattern,
                context_lines=int(args.get("context_lines") or 2),
                max_matches=int(args.get("max_matches") or 20),
                match_offset=max(0, int(args.get("match_offset") or 0)),
                heading_scope=heading_scope or None,
                full_body=body,
            )

        heading = (args.get("heading") or "").strip()
        if heading:
            slab = _slice_by_heading(body, heading)
            if slab is None:
                sheet_names = _list_sheet_headings(body)
                return SegmentResult(
                    error=f"sheet/heading not found: {heading!r}",
                    extras={"available_sheets": sheet_names},
                )
            return _clamp_ss(
                slab, offset, max_chars,
                extras={"heading": heading},
            )

        return _clamp_ss(body, offset, max_chars)

    @classmethod
    async def _extract_text(cls, storage: StorageBackend, key: str) -> str:
        body, _coverage = await cls._extract_text_with_coverage(storage, key)
        return body

    @classmethod
    async def _extract_read_text(cls, storage: StorageBackend, key: str) -> str:
        buf = bytearray()
        async for chunk in storage.get(key):
            buf.extend(chunk)
        return cls._render_read_from_bytes(bytes(buf))

    @classmethod
    async def _extract_text_with_coverage(
        cls, storage: StorageBackend, key: str,
    ) -> tuple[str, dict[str, Any]]:
        buf = bytearray()
        async for chunk in storage.get(key):
            buf.extend(chunk)
        # openpyxl workbook parsing is pure-CPU; offload it so the event loop
        # and worker heartbeats stay responsive on large spreadsheets.
        text, coverage = await asyncio.to_thread(
            cls._render_from_bytes_with_coverage, bytes(buf),
        )
        coverage["total_bytes"] = len(buf)
        coverage["indexed_bytes"] = len(buf)
        return text, coverage

    @staticmethod
    def _render_from_bytes(body: bytes) -> str:
        text, _coverage = SpreadsheetPipeline._render_from_bytes_with_coverage(body)
        return text

    @staticmethod
    def _render_read_from_bytes(body: bytes) -> str:
        text, _coverage = SpreadsheetPipeline._render_from_bytes_with_coverage(
            body,
            read_full=True,
        )
        return text

    @staticmethod
    def _render_from_bytes_with_coverage(
        body: bytes,
        *,
        read_full: bool = False,
    ) -> tuple[str, dict[str, Any]]:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise RuntimeError(
                "spreadsheet pipeline needs openpyxl; "
                "`pip install openpyxl`"
            ) from exc
        wb = openpyxl.load_workbook(
            io.BytesIO(body),
            data_only=True,
            read_only=True,
        )
        try:
            return _render_workbook(wb, read_full=read_full)
        finally:
            wb.close()


def _render_workbook(wb: Any, *, read_full: bool = False) -> tuple[str, dict[str, Any]]:
    parts: list[str] = []
    sheets: list[dict[str, Any]] = []
    total_rows_all = 0
    indexed_rows_all = 0
    any_partial = False
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"# Sheet: {sheet_name}")
        rows = list(
            _iter_rows(
                ws,
                None if read_full else MAX_ROWS_PER_SHEET + MAX_TAIL_PEEK,
            )
        )
        total_rows_estimate = max(int(getattr(ws, "max_row", 0) or 0), len(rows))
        indexed_rows = len(rows) if read_full else min(len(rows), MAX_ROWS_PER_SHEET)
        sheet_partial = False if read_full else (
            len(rows) > MAX_ROWS_PER_SHEET
            or total_rows_estimate > MAX_ROWS_PER_SHEET
        )
        total_rows_all += total_rows_estimate
        indexed_rows_all += indexed_rows
        any_partial = any_partial or sheet_partial
        sheets.append({
            "name": sheet_name,
            "total_rows": total_rows_estimate,
            "indexed_rows": indexed_rows,
            "indexed_partial": sheet_partial,
        })
        if not rows:
            parts.append("(empty sheet)")
            continue
        if read_full or len(rows) <= MAX_ROWS_PER_SHEET:
            for r in rows:
                parts.append(_format_row(
                    r,
                    max_cell_chars=None if read_full else MAX_CELL_CHARS,
                ))
        else:
            for r in rows[:MAX_ROWS_PER_SHEET]:
                parts.append(_format_row(r))
            omitted = max(0, total_rows_estimate - MAX_ROWS_PER_SHEET)
            parts.append(
                f"\n[...{omitted}+ rows omitted from index preview...]"
            )
        parts.append("")
    coverage = {
        "unit": "rows",
        "source_mode": "spreadsheet_full_text" if read_full else "spreadsheet_row_sample",
        "total_units": total_rows_all,
        "indexed_units": indexed_rows_all,
        "total_rows": total_rows_all,
        "indexed_rows": indexed_rows_all,
        "indexed_partial": any_partial,
        "partial_reasons": ["sheet_row_cap"] if any_partial else [],
        "sheet_count": len(sheets),
        "sheets": sheets[:50],
        "chunked": False,
        "chunk_count": 1,
        "text_truncated": any_partial,
    }
    if not read_full:
        coverage["max_rows_per_sheet"] = MAX_ROWS_PER_SHEET
    return "\n".join(parts).strip(), coverage


def _sheet_sections(coverage: dict[str, Any]) -> list[dict[str, Any]]:
    """Represent each indexed worksheet as a stable named section."""
    sections: list[dict[str, Any]] = []
    raw_sheets = coverage.get("sheets")
    if not isinstance(raw_sheets, list):
        return sections
    for index, sheet in enumerate(raw_sheets[:200], start=1):
        if not isinstance(sheet, dict):
            continue
        name = str(sheet.get("name") or f"Sheet {index}").strip()
        indexed_rows = max(0, int(sheet.get("indexed_rows") or 0))
        total_rows = max(indexed_rows, int(sheet.get("total_rows") or 0))
        partial = bool(sheet.get("indexed_partial"))
        scope = f"rows 1-{indexed_rows}" if indexed_rows else "empty sheet"
        if partial:
            scope += f" of approximately {total_rows}"
        sections.append({
            "id": f"s{index}",
            "title": name,
            "anchor": {"unit": "sheet", "value": name},
            "summary": scope,
            "key_terms": [name],
        })
    return sections


def _iter_rows(ws: Any, hard_limit: int | None):
    count = 0
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            yield row
            count += 1
            if hard_limit is not None and count >= hard_limit:
                return


def _format_row(row: tuple, *, max_cell_chars: int | None = MAX_CELL_CHARS) -> str:
    cells: list[str] = []
    for c in row:
        if c is None:
            cells.append("")
            continue
        s = str(c)
        if max_cell_chars is not None and len(s) > max_cell_chars:
            s = s[:max_cell_chars] + "..."
        cells.append(s.replace("|", r"\|").replace("\n", " "))
    return " | ".join(cells)


# ---- read_segment helpers --------------------------------------------------

_SHEET_HEADING_RX = re.compile(r"^# Sheet: (.+)$", re.MULTILINE)


def _list_sheet_headings(body: str) -> list[str]:
    return _SHEET_HEADING_RX.findall(body)


def _slice_by_heading(body: str, heading: str) -> str | None:
    """Return the body of the named sheet (everything from its `# Sheet: name`
    line up to the next `# Sheet:` line or EOF). The heading argument may
    be the full `Sheet: name` form or just `name`."""
    target = heading.removeprefix("Sheet: ").strip()
    matches = list(_SHEET_HEADING_RX.finditer(body))
    for i, m in enumerate(matches):
        if m.group(1).strip() == target:
            start = m.start()
            end = matches[i + 1].start() if i + 1 < len(matches) else len(body)
            return body[start:end].rstrip()
    return None


def _clamp_ss(
    text: str, offset: int, max_chars: int,
    *, extras: dict[str, Any] | None = None,
) -> SegmentResult:
    extras = dict(extras or {})
    total = len(text)
    chunk = text[offset: offset + max_chars]
    truncated = (offset + len(chunk)) < total
    extras.update({
        "offset": offset,
        "char_count": len(chunk),
        "total_chars": total,
        "truncated": truncated,
        "available_sheets": _list_sheet_headings(text),
    })
    if truncated:
        extras["next_offset"] = offset + len(chunk)
    if not chunk:
        return SegmentResult(text="", error="empty result", extras=extras)
    return SegmentResult(text=chunk, extras=extras)


def _ss_pattern_search(
    *, body: str, pattern: str, context_lines: int, max_matches: int,
    match_offset: int = 0,
    heading_scope: str | None = None,
    full_body: str | None = None,
) -> SegmentResult:
    try:
        rx = re.compile(pattern, re.IGNORECASE | re.MULTILINE)
    except re.error as exc:
        return SegmentResult(error=f"invalid regex: {exc}")

    sheets_source = full_body if full_body is not None else body

    lines = body.splitlines()
    all_hits: list[dict[str, Any]] = []
    current_sheet = heading_scope or ""
    for i, line in enumerate(lines, start=1):
        m_sheet = _SHEET_HEADING_RX.match(line)
        if m_sheet:
            current_sheet = m_sheet.group(1).strip()
            continue
        for m in rx.finditer(line):
            s = max(0, i - 1 - context_lines)
            e = min(len(lines), i + context_lines)
            all_hits.append({
                "sheet": current_sheet,
                "line": i,
                "match": m.group(0)[:200],
                "context": "\n".join(lines[s:e]),
            })

    total = len(all_hits)
    hits = all_hits[match_offset: match_offset + max_matches]
    has_more = (match_offset + len(hits)) < total

    extras: dict[str, Any] = {
        "pattern": pattern,
        "match_count": len(hits),
        "total_matches": total,
        "match_offset": match_offset,
        "has_more": has_more,
        "hits": hits,
        "available_sheets": _list_sheet_headings(sheets_source),
    }
    if heading_scope:
        extras["scope_heading"] = heading_scope
    if has_more:
        extras["next_match_offset"] = match_offset + len(hits)

    if not hits:
        if match_offset and total:
            err = f"match_offset {match_offset} exceeds total_matches {total}"
        else:
            err = "no matches"
        return SegmentResult(text="", error=err, extras=extras)

    rendered = "\n\n".join(
        f"[{h['sheet']} L{h['line']}] {h['match']}\n  ┊ {h['context']}"
        for h in hits
    )
    return SegmentResult(text=rendered, extras=extras)
